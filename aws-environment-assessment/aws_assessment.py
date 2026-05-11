#!/usr/bin/env python3
"""
AWS Environment Assessment Tool

Read-only inventory tool that scans an AWS account and produces a single
multi-sheet Excel workbook covering every major workload type — EC2, EBS,
RDS, S3, EFS, FSx, DynamoDB, Redshift, EKS, ECS, Lambda, WorkSpaces,
DocumentDB, ElastiCache, and AWS Backup.

DISCLAIMER
----------
This is a community sample script provided without support guarantees.
It is not an official product and is not covered by any support agreement.
Use at your own risk. Review the code before running it in any environment.

Requirements:
    pip install boto3 openpyxl tqdm

Usage:
    python aws_assessment.py
    python aws_assessment.py --regions us-east-1 us-west-2 --profile myprofile
    python aws_assessment.py --all-regions --output my_assessment.xlsx
"""

import boto3
import json
import sys
import argparse
import datetime
import logging
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

logging.basicConfig(level=logging.WARNING, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ─── Colour palette ───────────────────────────────────
C_HEADER_FILL   = "1A5276"   # dark navy
C_HEADER_FONT   = "FFFFFF"
C_SUBHDR_FILL   = "2E86C1"   # mid blue
C_ALT_ROW       = "EBF5FB"   # pale blue
C_WARN          = "F9E79F"   # yellow
C_CRITICAL      = "FADBD8"   # red-pink
C_GOOD          = "D5F5E3"   # green
C_SUMMARY_FILL  = "1E8449"   # green
C_SUMMARY_FONT  = "FFFFFF"

GiB = 1024 ** 3
TiB = 1024 ** 4

# ─── Helpers ──────────────────────────────────────────────────────────────────

MiB = 1024 ** 2

def gib(bytes_val):
    if bytes_val is None:
        return 0
    return round(bytes_val / GiB, 4)

def mib(bytes_val):
    if bytes_val is None:
        return 0
    return round(bytes_val / MiB, 1)

def tib(bytes_val):
    if bytes_val is None:
        return 0
    return round(bytes_val / TiB, 6)

def safe_get(d, *keys, default=""):
    for k in keys:
        if not isinstance(d, dict):
            return default
        d = d.get(k, default)
    return d if d is not None else default

def tag_value(tags, key):
    if not tags:
        return ""
    for t in tags:
        if t.get("Key", "").lower() == key.lower():
            return t.get("Value", "")
    return ""

def get_all_regions(session):
    ec2 = session.client("ec2", region_name="us-east-1")
    resp = ec2.describe_regions(AllRegions=False)
    return [r["RegionName"] for r in resp["Regions"]]

def make_client(session, service, region):
    return session.client(service, region_name=region)

def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def now_str():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

# ─── Data collectors ──────────────────────────────────────────────────────────

def collect_ec2(session, region):
    rows = []
    ec2 = make_client(session, "ec2", region)
    paginator = ec2.get_paginator("describe_instances")
    for page in paginator.paginate():
        for res in page["Reservations"]:
            for inst in res["Instances"]:
                iid   = inst["InstanceId"]
                state = inst["State"]["Name"]
                itype = inst.get("InstanceType", "")
                name  = tag_value(inst.get("Tags", []), "Name")
                env   = tag_value(inst.get("Tags", []), "Environment")
                owner = tag_value(inst.get("Tags", []), "Owner")
                os    = inst.get("Platform", "Linux/Unix")
                az    = inst.get("Placement", {}).get("AvailabilityZone", "")

                # Root volume
                root_size_gib = 0
                data_size_gib = 0
                vol_count     = 0
                vol_ids       = []

                for mapping in inst.get("BlockDeviceMappings", []):
                    vid = mapping.get("Ebs", {}).get("VolumeId", "")
                    vol_ids.append(vid)
                    vol_count += 1

                # Fetch actual volume sizes
                if vol_ids:
                    try:
                        vols = ec2.describe_volumes(VolumeIds=vol_ids)["Volumes"]
                        for v in vols:
                            sz = v.get("Size", 0)  # GiB
                            if v.get("Attachments") and v["Attachments"][0].get("Device") in ["/dev/sda1", "/dev/xvda", "xvda"]:
                                root_size_gib += sz
                            else:
                                data_size_gib += sz
                    except Exception:
                        pass

                total_gib = root_size_gib + data_size_gib

                # Backup plan
                backup_tag = tag_value(inst.get("Tags", []), "backup")
                has_backup = "Yes" if backup_tag.lower() in ("true", "yes", "enabled") else "Unknown"

                rows.append({
                    "Region":           region,
                    "Instance ID":      iid,
                    "Name":             name,
                    "State":            state,
                    "Instance Type":    itype,
                    "OS / Platform":    os,
                    "AZ":               az,
                    "Environment":      env,
                    "Owner":            owner,
                    "Volume Count":     vol_count,
                    "Root Disk (GiB)":  root_size_gib,
                    "Data Disks (GiB)": data_size_gib,
                    "Total Storage (GiB)": total_gib,
                    "VPC ID":           inst.get("VpcId", ""),
                    "Subnet ID":        inst.get("SubnetId", ""),
                    "AMI ID":           inst.get("ImageId", ""),
                    "Launch Time":      str(inst.get("LaunchTime", ""))[:10],
                    "Tag:Backup":       backup_tag,
                    "Notes":            "",
                })
    return rows


def collect_ebs_volumes(session, region):
    rows = []
    ec2 = make_client(session, "ec2", region)
    paginator = ec2.get_paginator("describe_volumes")
    for page in paginator.paginate():
        for v in page["Volumes"]:
            attached_to = ", ".join(
                a.get("InstanceId", "") for a in v.get("Attachments", [])
            )
            rows.append({
                "Region":        region,
                "Volume ID":     v["VolumeId"],
                "Name":          tag_value(v.get("Tags", []), "Name"),
                "State":         v["State"],
                "Type":          v["VolumeType"],
                "Size (GiB)":    v["Size"],
                "IOPS":          v.get("Iops", ""),
                "Throughput":    v.get("Throughput", ""),
                "Encrypted":     v.get("Encrypted", False),
                "Multi-Attach":  v.get("MultiAttachEnabled", False),
                "AZ":            v["AvailabilityZone"],
                "Attached To":   attached_to,
                "Snapshot ID":   v.get("SnapshotId", ""),
                "Created":       str(v.get("CreateTime", ""))[:10],
            })
    return rows


def collect_ebs_snapshots(session, region, account_id):
    rows = []
    ec2 = make_client(session, "ec2", region)
    paginator = ec2.get_paginator("describe_snapshots")
    for page in paginator.paginate(OwnerIds=[account_id]):
        for s in page["Snapshots"]:
            rows.append({
                "Region":        region,
                "Snapshot ID":   s["SnapshotId"],
                "Name":          tag_value(s.get("Tags", []), "Name"),
                "Volume ID":     s.get("VolumeId", ""),
                "State":         s["State"],
                "Size (GiB)":    s.get("VolumeSize", 0),
                "Encrypted":     s.get("Encrypted", False),
                "Description":   s.get("Description", ""),
                "Start Time":    str(s.get("StartTime", ""))[:10],
            })
    return rows


def collect_rds(session, region):
    rows = []
    rds = make_client(session, "rds", region)
    paginator = rds.get_paginator("describe_db_instances")
    for page in paginator.paginate():
        for db in page["DBInstances"]:
            rows.append({
                "Region":                region,
                "DB Identifier":         db["DBInstanceIdentifier"],
                "Engine":                db["Engine"],
                "Engine Version":        db.get("EngineVersion", ""),
                "Instance Class":        db["DBInstanceClass"],
                "Status":                db["DBInstanceStatus"],
                "Multi-AZ":             db.get("MultiAZ", False),
                "Storage Type":          db.get("StorageType", ""),
                "Allocated Storage (GiB)": db.get("AllocatedStorage", 0),
                "Max Allocated (GiB)":   db.get("MaxAllocatedStorage", ""),
                "Encrypted":             db.get("StorageEncrypted", False),
                "Backup Retention (days)": db.get("BackupRetentionPeriod", 0),
                "Automated Backups":     "Yes" if db.get("BackupRetentionPeriod", 0) > 0 else "No",
                "DB Cluster ID":         db.get("DBClusterIdentifier", ""),
                "VPC":                   safe_get(db, "DBSubnetGroup", "VpcId"),
                "AZ":                    db.get("AvailabilityZone", ""),
                "License Model":         db.get("LicenseModel", ""),
                "Public":                db.get("PubliclyAccessible", False),
                "Created":               str(db.get("InstanceCreateTime", ""))[:10],
                "Notes":                 "",
            })

    # Aurora clusters
    try:
        cpaginator = rds.get_paginator("describe_db_clusters")
        for page in cpaginator.paginate():
            for cl in page["DBClusters"]:
                if cl.get("Engine", "").startswith("aurora"):
                    rows.append({
                        "Region":                region,
                        "DB Identifier":         cl["DBClusterIdentifier"] + " [CLUSTER]",
                        "Engine":                cl["Engine"],
                        "Engine Version":        cl.get("EngineVersion", ""),
                        "Instance Class":        "N/A (cluster)",
                        "Status":                cl["Status"],
                        "Multi-AZ":             len(cl.get("AvailabilityZones", [])) > 1,
                        "Storage Type":          "aurora",
                        "Allocated Storage (GiB)": cl.get("AllocatedStorage", 0),
                        "Max Allocated (GiB)":   "",
                        "Encrypted":             cl.get("StorageEncrypted", False),
                        "Backup Retention (days)": cl.get("BackupRetentionPeriod", 0),
                        "Automated Backups":     "Yes" if cl.get("BackupRetentionPeriod", 0) > 0 else "No",
                        "DB Cluster ID":         cl["DBClusterIdentifier"],
                        "VPC":                   cl.get("VpcId", ""),
                        "AZ":                    ", ".join(cl.get("AvailabilityZones", [])),
                        "License Model":         "",
                        "Public":                "",
                        "Created":               str(cl.get("ClusterCreateTime", ""))[:10],
                        "Notes":                 f"Members: {len(cl.get('DBClusterMembers', []))}",
                    })
    except Exception:
        pass

    return rows


def collect_s3(session):
    rows = []
    s3  = session.client("s3")
    cw  = session.client("cloudwatch", region_name="us-east-1")
    try:
        buckets = s3.list_buckets().get("Buckets", [])
    except Exception as e:
        log.warning("S3 list_buckets failed: %s", e)
        return rows

    for b in buckets:
        name = b["Name"]
        created = str(b.get("CreationDate", ""))[:10]

        # Region
        try:
            loc = s3.get_bucket_location(Bucket=name)
            region = loc.get("LocationConstraint") or "us-east-1"
        except Exception:
            region = "unknown"

        # Size: CloudWatch first (accurate for large buckets, updates daily),
        # fall back to direct listing (always current, slow on huge buckets).
        size_bytes   = None
        object_count = None
        try:
            end   = datetime.datetime.now(datetime.timezone.utc)
            start = end - datetime.timedelta(days=3)
            for storage_type in ("StandardStorage", "AllStorageTypes"):
                resp = cw.get_metric_statistics(
                    Namespace="AWS/S3",
                    MetricName="BucketSizeBytes",
                    Dimensions=[
                        {"Name": "BucketName",  "Value": name},
                        {"Name": "StorageType", "Value": storage_type},
                    ],
                    StartTime=start,
                    EndTime=end,
                    Period=86400,
                    Statistics=["Average"],
                )
                pts = sorted(resp["Datapoints"], key=lambda x: x["Timestamp"])
                if pts:
                    size_bytes = pts[-1]["Average"]
                    break

            resp2 = cw.get_metric_statistics(
                Namespace="AWS/S3",
                MetricName="NumberOfObjects",
                Dimensions=[
                    {"Name": "BucketName",  "Value": name},
                    {"Name": "StorageType", "Value": "AllStorageTypes"},
                ],
                StartTime=start,
                EndTime=end,
                Period=86400,
                Statistics=["Average"],
            )
            pts2 = sorted(resp2["Datapoints"], key=lambda x: x["Timestamp"])
            if pts2:
                object_count = int(pts2[-1]["Average"])
        except Exception:
            pass

        # Fallback: direct listing (exact, works immediately, capped at 100k objects)
        if size_bytes is None:
            try:
                total_bytes = 0
                total_objs  = 0
                paginator = s3.get_paginator("list_objects_v2")
                for page in paginator.paginate(Bucket=name, PaginationConfig={"MaxItems": 100_000}):
                    for obj in page.get("Contents", []):
                        total_bytes += obj.get("Size", 0)
                        total_objs  += 1
                size_bytes   = total_bytes
                object_count = total_objs
            except Exception:
                pass

        # Versioning
        versioning = ""
        try:
            vr = s3.get_bucket_versioning(Bucket=name)
            versioning = vr.get("Status", "Disabled")
        except Exception:
            pass

        # Lifecycle
        lifecycle = "None"
        try:
            lc = s3.get_bucket_lifecycle_configuration(Bucket=name)
            lifecycle = f"{len(lc.get('Rules', []))} rules"
        except Exception:
            pass

        # Replication
        replication = "No"
        try:
            rr = s3.get_bucket_replication(Bucket=name)
            if rr.get("ReplicationConfiguration"):
                replication = "Yes"
        except Exception:
            pass

        # Encryption
        encryption = "None"
        try:
            enc = s3.get_bucket_encryption(Bucket=name)
            rules = enc.get("ServerSideEncryptionConfiguration", {}).get("Rules", [])
            if rules:
                algo = rules[0].get("ApplyServerSideEncryptionByDefault", {}).get("SSEAlgorithm", "")
                encryption = algo
        except Exception:
            pass

        # Public access
        public_access = "Unknown"
        try:
            pa = s3.get_public_access_block(Bucket=name)
            cfg = pa.get("PublicAccessBlockConfiguration", {})
            public_access = "Blocked" if all(cfg.values()) else "Partial/Open"
        except Exception:
            pass

        rows.append({
            "Region":           region,
            "Bucket Name":      name,
            "Created":          created,
            "Size (MiB)":       mib(size_bytes) if size_bytes is not None else "N/A",
            "Size (GiB)":       gib(size_bytes) if size_bytes is not None else "N/A",
            "Size (TiB)":       tib(size_bytes) if size_bytes is not None else "N/A",
            "Object Count":     object_count or "N/A",
            "Versioning":       versioning,
            "Replication":      replication,
            "Lifecycle Rules":  lifecycle,
            "Encryption":       encryption,
            "Public Access":    public_access,
            "Notes":            "",
        })

    return rows


def collect_efs(session, region):
    rows = []
    efs = make_client(session, "efs", region)
    try:
        paginator = efs.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page["FileSystems"]:
                sz = fs.get("SizeInBytes", {})
                rows.append({
                    "Region":               region,
                    "File System ID":       fs["FileSystemId"],
                    "Name":                 tag_value(fs.get("Tags", []), "Name"),
                    "State":                fs["LifeCycleState"],
                    "Performance Mode":     fs.get("PerformanceMode", ""),
                    "Throughput Mode":      fs.get("ThroughputMode", ""),
                    "Provisioned Throughput (MiBps)": fs.get("ProvisionedThroughputInMibps", ""),
                    "Encrypted":            fs.get("Encrypted", False),
                    "Size (GiB)":           gib(sz.get("Value", 0)),
                    "IA Size (GiB)":        gib(sz.get("ValueInIA", 0)),
                    "Standard Size (GiB)":  gib(sz.get("ValueInStandard", 0)),
                    "Mount Targets":        "",
                    "Backup Policy":        "",
                    "Created":              str(fs.get("CreationTime", ""))[:10],
                })
    except Exception as e:
        log.warning("EFS %s: %s", region, e)
    return rows


def collect_fsx(session, region):
    rows = []
    try:
        fsx = make_client(session, "fsx", region)
        paginator = fsx.get_paginator("describe_file_systems")
        for page in paginator.paginate():
            for fs in page["FileSystems"]:
                ftype = fs.get("FileSystemType", "")
                cap   = fs.get("StorageCapacity", 0)
                name  = tag_value(fs.get("Tags", []), "Name")

                extra = ""
                if ftype == "WINDOWS":
                    wd = fs.get("WindowsConfiguration", {})
                    extra = f"Throughput: {wd.get('ThroughputCapacity','')} MBps | HA: {wd.get('DeploymentType','')}"
                elif ftype == "LUSTRE":
                    ld = fs.get("LustreConfiguration", {})
                    extra = f"Deployment: {ld.get('DeploymentType','')} | Per-unit: {ld.get('PerUnitStorageThroughput','')} MBps"
                elif ftype == "ONTAP":
                    od = fs.get("OntapConfiguration", {})
                    extra = f"Deployment: {od.get('DeploymentType','')} | Throughput: {od.get('ThroughputCapacity','')} MBps"
                elif ftype == "OPENZFS":
                    zd = fs.get("OpenZFSConfiguration", {})
                    extra = f"Deployment: {zd.get('DeploymentType','')} | Throughput: {zd.get('ThroughputCapacity','')} MBps"

                rows.append({
                    "Region":               region,
                    "File System ID":       fs["FileSystemId"],
                    "Name":                 name,
                    "Type":                 ftype,
                    "State":                fs["Lifecycle"],
                    "Storage Type":         fs.get("StorageType", ""),
                    "Capacity (GiB)":       cap,
                    "Encrypted":            fs.get("KmsKeyId", "") != "",
                    "VPC":                  fs.get("VpcId", ""),
                    "AZs":                  ", ".join(fs.get("SubnetIds", [])),
                    "Configuration":        extra,
                    "Created":              str(fs.get("CreationTime", ""))[:10],
                })
    except Exception as e:
        log.warning("FSx %s: %s", region, e)
    return rows


def collect_dynamodb(session, region):
    rows = []
    try:
        ddb = make_client(session, "dynamodb", region)
        paginator = ddb.get_paginator("list_tables")
        for page in paginator.paginate():
            for tname in page["TableNames"]:
                try:
                    t = ddb.describe_table(TableName=tname)["Table"]
                    sz   = t.get("TableSizeBytes", 0)
                    items = t.get("ItemCount", 0)
                    billing = t.get("BillingModeSummary", {}).get("BillingMode", "PROVISIONED")
                    pib = t.get("ProvisionedThroughput", {})
                    backup = t.get("ContinuousBackupsDescription", {})

                    pitr = "Unknown"
                    try:
                        cb = ddb.describe_continuous_backups(TableName=tname)
                        pitr_status = cb.get("ContinuousBackupsDescription", {}) \
                                        .get("PointInTimeRecoveryDescription", {}) \
                                        .get("PointInTimeRecoveryStatus", "DISABLED")
                        pitr = pitr_status
                    except Exception:
                        pass

                    rows.append({
                        "Region":           region,
                        "Table Name":       tname,
                        "Status":           t.get("TableStatus", ""),
                        "Billing Mode":     billing,
                        "Size (GiB)":       gib(sz),
                        "Item Count":       items,
                        "RCU":              pib.get("ReadCapacityUnits", "on-demand"),
                        "WCU":              pib.get("WriteCapacityUnits", "on-demand"),
                        "PITR Enabled":     pitr,
                        "Global Tables":    len(t.get("Replicas", [])) > 0,
                        "Streams":          t.get("StreamSpecification", {}).get("StreamEnabled", False),
                        "Encrypted":        t.get("SSEDescription", {}).get("Status", "DISABLED") == "ENABLED",
                        "Created":          str(t.get("CreationDateTime", ""))[:10],
                    })
                except Exception:
                    pass
    except Exception as e:
        log.warning("DynamoDB %s: %s", region, e)
    return rows


def collect_redshift(session, region):
    rows = []
    try:
        rs = make_client(session, "redshift", region)
        paginator = rs.get_paginator("describe_clusters")
        for page in paginator.paginate():
            for cl in page["Clusters"]:
                nodes      = cl.get("NumberOfNodes", 1)
                node_type  = cl.get("NodeType", "")
                # Approximate storage — actual size varies by node type
                rows.append({
                    "Region":               region,
                    "Cluster ID":           cl["ClusterIdentifier"],
                    "Status":               cl["ClusterStatus"],
                    "Node Type":            node_type,
                    "Nodes":                nodes,
                    "DB Name":              cl.get("DBName", ""),
                    "Encrypted":            cl.get("Encrypted", False),
                    "Backup Retention (days)": cl.get("AutomatedSnapshotRetentionPeriod", 0),
                    "Automated Backups":    "Yes" if cl.get("AutomatedSnapshotRetentionPeriod", 0) > 0 else "No",
                    "Public":               cl.get("PubliclyAccessible", False),
                    "VPC":                  cl.get("VpcId", ""),
                    "AZ":                   cl.get("AvailabilityZone", ""),
                    "Serverless":           False,
                    "Created":              str(cl.get("ClusterCreateTime", ""))[:10],
                })
    except Exception as e:
        log.warning("Redshift %s: %s", region, e)

    # Redshift Serverless
    try:
        rss = make_client(session, "redshift-serverless", region)
        ns_paginator = rss.get_paginator("list_namespaces")
        for page in ns_paginator.paginate():
            for ns in page["namespaces"]:
                rows.append({
                    "Region":               region,
                    "Cluster ID":           ns.get("namespaceName", "") + " [SERVERLESS]",
                    "Status":               ns.get("status", ""),
                    "Node Type":            "serverless",
                    "Nodes":                "N/A",
                    "DB Name":              ns.get("dbName", ""),
                    "Encrypted":            ns.get("kmsKeyId", "") != "",
                    "Backup Retention (days)": "",
                    "Automated Backups":    "",
                    "Public":               "",
                    "VPC":                  "",
                    "AZ":                   "",
                    "Serverless":           True,
                    "Created":              str(ns.get("creationDate", ""))[:10],
                })
    except Exception:
        pass

    return rows


def collect_eks(session, region):
    rows = []
    try:
        eks = make_client(session, "eks", region)
        clusters = eks.list_clusters().get("clusters", [])
        for cname in clusters:
            cl = eks.describe_cluster(name=cname)["cluster"]

            # Node groups
            ng_names = eks.list_nodegroups(clusterName=cname).get("nodegroups", [])
            total_nodes  = 0
            node_details = []
            for ng in ng_names:
                try:
                    ngd = eks.describe_nodegroup(clusterName=cname, nodegroupName=ng)["nodegroup"]
                    desired = ngd.get("scalingConfig", {}).get("desiredSize", 0)
                    total_nodes += desired
                    node_details.append(
                        f"{ng}({ngd.get('instanceTypes',['?'])[0]}×{desired})"
                    )
                except Exception:
                    pass

            rows.append({
                "Region":           region,
                "Cluster Name":     cname,
                "Status":           cl.get("status", ""),
                "K8s Version":      cl.get("version", ""),
                "Platform Version": cl.get("platformVersion", ""),
                "Node Groups":      len(ng_names),
                "Total Nodes":      total_nodes,
                "Node Details":     " | ".join(node_details),
                "VPC":              safe_get(cl, "resourcesVpcConfig", "vpcId"),
                "Private Endpoint": safe_get(cl, "resourcesVpcConfig", "endpointPrivateAccess"),
                "Logging":          ", ".join(
                    [l["type"] for l in cl.get("logging", {}).get("clusterLogging", [])
                     if l.get("enabled")]
                ),
                "Created":          str(cl.get("createdAt", ""))[:10],
            })
    except Exception as e:
        log.warning("EKS %s: %s", region, e)
    return rows


def collect_ecs(session, region):
    rows = []
    try:
        ecs = make_client(session, "ecs", region)
        cluster_arns = []
        paginator = ecs.get_paginator("list_clusters")
        for page in paginator.paginate():
            cluster_arns.extend(page["clusterArns"])

        for arn_chunk in chunks(cluster_arns, 100):
            clusters = ecs.describe_clusters(clusters=arn_chunk, include=["STATISTICS"])["clusters"]
            for cl in clusters:
                cname = cl["clusterName"]
                rows.append({
                    "Region":               region,
                    "Cluster Name":         cname,
                    "Status":               cl.get("status", ""),
                    "Active Services":      cl.get("activeServicesCount", 0),
                    "Running Tasks":        cl.get("runningTasksCount", 0),
                    "Pending Tasks":        cl.get("pendingTasksCount", 0),
                    "Registered Instances": cl.get("registeredContainerInstancesCount", 0),
                    "Capacity Providers":   ", ".join(cl.get("capacityProviders", [])),
                    "Notes":                "",
                })
    except Exception as e:
        log.warning("ECS %s: %s", region, e)
    return rows


def collect_lambda(session, region):
    rows = []
    try:
        lm = make_client(session, "lambda", region)
        paginator = lm.get_paginator("list_functions")
        for page in paginator.paginate():
            for fn in page["Functions"]:
                rows.append({
                    "Region":           region,
                    "Function Name":    fn["FunctionName"],
                    "Runtime":          fn.get("Runtime", ""),
                    "Memory (MB)":      fn.get("MemorySize", 128),
                    "Timeout (sec)":    fn.get("Timeout", 3),
                    "Package Size (MB)": round(fn.get("CodeSize", 0) / (1024*1024), 2),
                    "Architecture":     ", ".join(fn.get("Architectures", ["x86_64"])),
                    "Last Modified":    str(fn.get("LastModified", ""))[:10],
                    "Handler":          fn.get("Handler", ""),
                    "Description":      fn.get("Description", ""),
                })
    except Exception as e:
        log.warning("Lambda %s: %s", region, e)
    return rows


def collect_workspaces(session, region):
    rows = []
    try:
        ws = make_client(session, "workspaces", region)
        paginator = ws.get_paginator("describe_workspaces")
        for page in paginator.paginate():
            for w in page["Workspaces"]:
                rows.append({
                    "Region":           region,
                    "Workspace ID":     w["WorkspaceId"],
                    "User":             w.get("UserName", ""),
                    "State":            w.get("State", ""),
                    "Bundle ID":        w.get("BundleId", ""),
                    "Directory ID":     w.get("DirectoryId", ""),
                    "Running Mode":     w.get("WorkspaceProperties", {}).get("RunningMode", ""),
                    "Root Volume (GiB)": w.get("WorkspaceProperties", {}).get("RootVolumeSizeGib", ""),
                    "User Volume (GiB)": w.get("WorkspaceProperties", {}).get("UserVolumeSizeGib", ""),
                    "Compute Type":     w.get("WorkspaceProperties", {}).get("ComputeTypeName", ""),
                    "Protocol":         ", ".join(w.get("WorkspaceProperties", {}).get("Protocols", [])),
                })
    except Exception as e:
        log.warning("WorkSpaces %s: %s", region, e)
    return rows


def collect_documentdb(session, region):
    rows = []
    try:
        ddb = make_client(session, "docdb", region)
        paginator = ddb.get_paginator("describe_db_clusters")
        for page in paginator.paginate(Filters=[{"Name": "engine", "Values": ["docdb"]}]):
            for cl in page["DBClusters"]:
                rows.append({
                    "Region":               region,
                    "Cluster ID":           cl["DBClusterIdentifier"],
                    "Engine":               cl.get("Engine", "docdb"),
                    "Engine Version":       cl.get("EngineVersion", ""),
                    "Status":               cl["Status"],
                    "Members":              len(cl.get("DBClusterMembers", [])),
                    "Storage (GiB)":        cl.get("AllocatedStorage", 0),
                    "Encrypted":            cl.get("StorageEncrypted", False),
                    "Backup Retention (days)": cl.get("BackupRetentionPeriod", 0),
                    "Multi-AZ":             len(cl.get("AvailabilityZones", [])) > 1,
                    "VPC":                  cl.get("VpcId", ""),
                    "Created":              str(cl.get("ClusterCreateTime", ""))[:10],
                })
    except Exception as e:
        log.warning("DocumentDB %s: %s", region, e)
    return rows


def collect_elasticache(session, region):
    rows = []
    try:
        ec = make_client(session, "elasticache", region)
        paginator = ec.get_paginator("describe_cache_clusters")
        for page in paginator.paginate(ShowCacheNodeInfo=True):
            for cl in page["CacheClusters"]:
                rows.append({
                    "Region":           region,
                    "Cluster ID":       cl["CacheClusterId"],
                    "Engine":           cl.get("Engine", ""),
                    "Engine Version":   cl.get("EngineVersion", ""),
                    "Node Type":        cl.get("CacheNodeType", ""),
                    "Status":           cl.get("CacheClusterStatus", ""),
                    "Nodes":            cl.get("NumCacheNodes", 0),
                    "AZ":               cl.get("PreferredAvailabilityZone", ""),
                    "Replication Group": cl.get("ReplicationGroupId", ""),
                    "Encrypted at Rest": cl.get("AtRestEncryptionEnabled", False),
                    "Encrypted in Transit": cl.get("TransitEncryptionEnabled", False),
                    "Backup Retention (days)": cl.get("SnapshotRetentionLimit", 0),
                    "Created":          str(cl.get("CacheClusterCreateTime", ""))[:10],
                })
    except Exception as e:
        log.warning("ElastiCache %s: %s", region, e)
    return rows


def collect_aws_backup(session, region):
    """Collect AWS Backup vaults and recent jobs to assess backup coverage."""
    rows = []
    try:
        bk = make_client(session, "backup", region)
        paginator = bk.get_paginator("list_backup_vaults")
        for page in paginator.paginate():
            for vault in page["BackupVaultList"]:
                rows.append({
                    "Region":           region,
                    "Vault Name":       vault["BackupVaultName"],
                    "Recovery Points":  vault.get("NumberOfRecoveryPoints", 0),
                    "Encrypted":        vault.get("EncryptionKeyArn", "") != "",
                    "Locked":           vault.get("Locked", False),
                    "Min Retention (days)": vault.get("MinRetentionDays", ""),
                    "Max Retention (days)": vault.get("MaxRetentionDays", ""),
                    "Created":          str(vault.get("CreationDate", ""))[:10],
                })
    except Exception as e:
        log.warning("Backup vaults %s: %s", region, e)
    return rows


def collect_backup_plans(session, region):
    rows = []
    try:
        bk = make_client(session, "backup", region)
        paginator = bk.get_paginator("list_backup_plans")
        for page in paginator.paginate():
            for plan in page["BackupPlansList"]:
                try:
                    detail = bk.get_backup_plan(BackupPlanId=plan["BackupPlanId"])["BackupPlan"]
                    for rule in detail.get("Rules", []):
                        rows.append({
                            "Region":           region,
                            "Plan Name":        plan["BackupPlanName"],
                            "Rule Name":        rule.get("RuleName", ""),
                            "Target Vault":     rule.get("TargetBackupVaultName", ""),
                            "Schedule":         rule.get("ScheduleExpression", ""),
                            "Start Window (min)": rule.get("StartWindowMinutes", ""),
                            "Completion Window (min)": rule.get("CompletionWindowMinutes", ""),
                            "Delete After (days)": safe_get(rule, "Lifecycle", "DeleteAfterDays"),
                            "Cold After (days)": safe_get(rule, "Lifecycle", "MoveToColdStorageAfterDays"),
                            "Copy To Region":   ", ".join(
                                c.get("DestinationBackupVaultArn", "").split(":")[3]
                                for c in rule.get("CopyActions", [])
                            ),
                            "Created":          str(plan.get("CreationDate", ""))[:10],
                        })
                except Exception:
                    pass
    except Exception as e:
        log.warning("Backup plans %s: %s", region, e)
    return rows


def get_account_id(session):
    try:
        return session.client("sts").get_caller_identity()["Account"]
    except Exception:
        return "unknown"


# ─── Excel builder ────────────────────────────────────────────────────────────

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font(hex_color=C_HEADER_FONT, bold=True, size=10):
    return Font(color=hex_color, bold=bold, size=size)

def apply_header_row(ws, headers, row=1, fill_color=C_HEADER_FILL, font_color=C_HEADER_FONT):
    fill = hex_fill(fill_color)
    font = header_font(font_color)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def auto_col_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        best = min_w
        for cell in col:
            if cell.value is not None:
                best = max(best, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[col_letter].width = best

def freeze_top_row(ws):
    ws.freeze_panes = ws.cell(row=2, column=1)

def add_sheet(wb, name, rows, columns, color_row=None, title=None):
    """Generic sheet writer."""
    ws = wb.create_sheet(title=name[:31])
    ws.sheet_properties.tabColor = C_SUMMARY_FILL

    start_row = 1
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13, color=C_SUMMARY_FILL)
        start_row = 2

    apply_header_row(ws, columns, row=start_row)

    for i, row in enumerate(rows):
        r = start_row + 1 + i
        fill = hex_fill(C_ALT_ROW) if i % 2 == 0 else None
        for j, col in enumerate(columns):
            cell = ws.cell(row=r, column=j+1, value=row.get(col, ""))
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if color_row:
                color_row(cell, col, row)

    freeze_top_row(ws)
    auto_col_width(ws)
    ws.row_dimensions[start_row].height = 30
    return ws


def build_summary_sheet(wb, data, account_id, regions, assessed_at):  # noqa: C901
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = C_SUMMARY_FILL

    # ── helpers ───────────────────────────────────────────────────────────────
    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def cell(r, c, value="", bold=False, size=10, color="000000",
             bg=None, align="left", wrap=False, italic=False):
        cl = ws.cell(row=r, column=c, value=value)
        cl.font = Font(bold=bold, size=size, color=color, italic=italic)
        cl.alignment = Alignment(horizontal=align, vertical="center",
                                 wrap_text=wrap)
        if bg:
            cl.fill = hex_fill(bg)
        return cl

    def section_header(r, c1, c2, label):
        merge(r, c1, r, c2)
        cl = cell(r, c1, label, bold=True, size=11, color=C_SUMMARY_FONT,
                  bg=C_HEADER_FILL, align="left")
        cl.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=1)
        ws.row_dimensions[r].height = 22
        return r + 1

    def kpi_tile(r, c, label, value, sub="", bg=C_SUBHDR_FILL, fg=C_SUMMARY_FONT):
        merge(r, c, r, c + 1)
        cell(r, c, label, bold=False, size=9, color="AED6F1",
             bg=bg, align="center")
        merge(r + 1, c, r + 1, c + 1)
        cell(r + 1, c, value, bold=True, size=20, color=fg,
             bg=bg, align="center")
        merge(r + 2, c, r + 2, c + 1)
        cell(r + 2, c, sub, bold=False, size=8, color="AED6F1",
             bg=bg, align="center")
        ws.row_dimensions[r].height     = 16
        ws.row_dimensions[r + 1].height = 30
        ws.row_dimensions[r + 2].height = 14

    def thin_border():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── pre-compute all data ──────────────────────────────────────────────────
    ec2_rows  = data.get("EC2 Instances",    [])
    ebs_rows  = data.get("EBS Volumes",      [])
    snap_rows = data.get("EBS Snapshots",    [])
    rds_rows  = data.get("RDS & Aurora",     [])
    s3_rows   = data.get("S3 Buckets",       [])
    efs_rows  = data.get("EFS",              [])
    fsx_rows  = data.get("FSx",              [])
    ddb_rows  = data.get("DynamoDB",         [])
    rs_rows   = data.get("Redshift",         [])
    eks_rows  = data.get("EKS",              [])
    ecs_rows  = data.get("ECS",              [])
    lam_rows  = data.get("Lambda",           [])
    ws2_rows  = data.get("WorkSpaces",       [])
    doc_rows  = data.get("DocumentDB",       [])
    ec_rows   = data.get("ElastiCache",      [])
    bkv_rows  = data.get("AWS Backup Vaults",[])
    bkp_rows  = data.get("AWS Backup Plans", [])

    def gib_sum(rows, key):
        return round(sum(r.get(key, 0) or 0 for r in rows), 2)

    def pct_enc(rows, key):
        if not rows:
            return "—"
        n = sum(1 for r in rows
                if r.get(key) in (True, "True", "true", "AES256", "aws:kms",
                                  "Yes", "yes", "ENABLED"))
        return f"{round(n / len(rows) * 100)}%"

    def count_state(rows, key, val):
        return sum(1 for r in rows if str(r.get(key, "")).lower() == val.lower())

    s3_gib   = round(sum(r["Size (GiB)"] for r in s3_rows
                         if isinstance(r.get("Size (GiB)"), (int, float))), 2)
    s3_mib   = round(sum(r["Size (MiB)"] for r in s3_rows
                         if isinstance(r.get("Size (MiB)"), (int, float))), 1)
    s3_objs  = sum(r.get("Object Count", 0) or 0 for r in s3_rows
                   if isinstance(r.get("Object Count"), int))

    ec2_gib  = gib_sum(ec2_rows,  "Total Storage (GiB)")
    ebs_gib  = gib_sum(ebs_rows,  "Size (GiB)")
    rds_gib  = gib_sum(rds_rows,  "Allocated Storage (GiB)")
    efs_gib  = gib_sum(efs_rows,  "Size (GiB)")
    fsx_gib  = gib_sum(fsx_rows,  "Capacity (GiB)")
    ddb_gib  = gib_sum(ddb_rows,  "Size (GiB)")
    ws2_gib  = round(sum(
        (r.get("Root Volume (GiB)", 0) or 0) +
        (r.get("User Volume (GiB)", 0) or 0)
        for r in ws2_rows), 2)

    total_gib  = round(ec2_gib + ebs_gib + rds_gib + s3_gib + efs_gib +
                       fsx_gib + ddb_gib + ws2_gib, 2)
    total_tib  = round(total_gib / 1024, 3)
    total_res  = (len(ec2_rows) + len(rds_rows) + len(s3_rows) + len(efs_rows) +
                  len(fsx_rows) + len(ddb_rows) + len(rs_rows) + len(eks_rows) +
                  len(ecs_rows) + len(lam_rows) + len(ws2_rows) + len(doc_rows) +
                  len(ec_rows))

    ec2_running = count_state(ec2_rows, "State", "running")
    ec2_stopped = count_state(ec2_rows, "State", "stopped")
    rds_avail   = count_state(rds_rows, "Status", "available")
    rds_stopped = count_state(rds_rows, "Status", "stopped")

    snap_gib = gib_sum(snap_rows, "Size (GiB)")

    # Risk findings
    unenc_ebs   = [r for r in ebs_rows if not r.get("Encrypted")]
    unatt_ebs   = [r for r in ebs_rows if r.get("State") == "available"]
    no_backup_ec2 = [r for r in ec2_rows
                     if not r.get("Tag:Backup") or
                     str(r.get("Tag:Backup", "")).lower()
                     not in ("true", "yes", "enabled")]
    rds_no_bk   = [r for r in rds_rows if r.get("Automated Backups") == "No"]
    pub_s3      = [r for r in s3_rows
                   if r.get("Public Access") not in ("Blocked", "")]
    no_vers_s3  = [r for r in s3_rows
                   if r.get("Versioning") in ("Disabled", "Suspended", "")]
    ddb_no_pitr = [r for r in ddb_rows if r.get("PITR Enabled") == "DISABLED"]
    pub_rds     = [r for r in rds_rows if r.get("Public") is True]

    # Region distribution
    from collections import Counter
    region_counts = Counter()
    for dataset in [ec2_rows, ebs_rows, rds_rows, efs_rows, fsx_rows,
                    ddb_rows, eks_rows, ecs_rows, lam_rows, ws2_rows]:
        for r in dataset:
            if r.get("Region"):
                region_counts[r["Region"]] += 1
    for r in s3_rows:
        if r.get("Region"):
            region_counts[r["Region"]] += 1

    # Backup vault summary
    total_rp    = sum(r.get("Recovery Points", 0) or 0 for r in bkv_rows)
    locked_vaults = sum(1 for r in bkv_rows if r.get("Locked"))

    # ── layout: 12 columns, rows 1–N ─────────────────────────────────────────
    COLS = 12

    # ── Row 1-2: Title banner ─────────────────────────────────────────────────
    merge(1, 1, 2, COLS)
    c = ws.cell(row=1, column=1)
    c.value = "AWS Environment Assessment"
    c.font  = Font(bold=True, size=22, color=C_SUMMARY_FONT)
    c.fill  = hex_fill(C_SUMMARY_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 14

    # ── Row 3: Meta bar ───────────────────────────────────────────────────────
    merge(3, 1, 3, COLS)
    c = ws.cell(row=3, column=1)
    c.value = (f"Account ID: {account_id}   |   "
               f"Regions Scanned: {len(regions)}  ({', '.join(regions)})   |   "
               f"Generated: {assessed_at}")
    c.font      = Font(size=9, italic=True, color="FFFFFF")
    c.fill      = hex_fill("2471A3")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 16

    ws.row_dimensions[4].height = 8   # spacer

    # ── Rows 5-7: KPI tiles (cols 1-12, pairs of 2) ──────────────────────────
    kpi_tile(5, 1,  "TOTAL RESOURCES",   str(total_res),   "across all services")
    kpi_tile(5, 3,  "TOTAL STORAGE",     f"{total_tib} TiB", f"{total_gib} GiB")
    kpi_tile(5, 5,  "EC2 INSTANCES",     str(len(ec2_rows)),
             f"{ec2_running} running  /  {ec2_stopped} stopped")
    kpi_tile(5, 7,  "RDS / AURORA",      str(len(rds_rows)),
             f"{rds_avail} available  /  {rds_stopped} stopped")
    kpi_tile(5, 9,  "S3 BUCKETS",        str(len(s3_rows)),
             f"{s3_mib} MiB  |  {s3_objs:,} objects")
    kpi_tile(5, 11, "EBS SNAPSHOTS",     str(len(snap_rows)),
             f"{snap_gib} GiB total")

    ws.row_dimensions[8].height = 8   # spacer

    # ── Row 9+: Two-column layout ─────────────────────────────────────────────
    # LEFT  = cols 1-6    RIGHT = cols 7-12
    cur_l = 9   # current left row
    cur_r = 9   # current right row

    # ── LEFT: Workload Inventory ──────────────────────────────────────────────
    cur_l = section_header(cur_l, 1, 6, "  Workload Inventory")

    inv_hdrs = ["Workload", "Count", "Storage (GiB)", "Storage (TiB)", "Encrypted %", "Regions"]
    for ci, h in enumerate(inv_hdrs, 1):
        c = ws.cell(row=cur_l, column=ci, value=h)
        c.fill = hex_fill(C_SUBHDR_FILL)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[cur_l].height = 20
    cur_l += 1

    def inv_regions(rows):
        regs = set(r.get("Region", "") for r in rows if r.get("Region"))
        return ", ".join(sorted(regs)) if regs else "—"

    inventory = [
        ("EC2 Instances",    len(ec2_rows),  ec2_gib,  round(ec2_gib/1024,4),   pct_enc(ebs_rows,"Encrypted"),  inv_regions(ec2_rows)),
        ("EBS Volumes",      len(ebs_rows),  ebs_gib,  round(ebs_gib/1024,4),   pct_enc(ebs_rows,"Encrypted"),  inv_regions(ebs_rows)),
        ("RDS / Aurora",     len(rds_rows),  rds_gib,  round(rds_gib/1024,4),   pct_enc(rds_rows,"Encrypted"),  inv_regions(rds_rows)),
        ("S3 Buckets",       len(s3_rows),   s3_gib,   round(s3_gib/1024,4),    pct_enc(s3_rows,"Encryption"),  inv_regions(s3_rows)),
        ("EFS",              len(efs_rows),  efs_gib,  round(efs_gib/1024,4),   pct_enc(efs_rows,"Encrypted"),  inv_regions(efs_rows)),
        ("FSx",              len(fsx_rows),  fsx_gib,  round(fsx_gib/1024,4),   "—",                            inv_regions(fsx_rows)),
        ("DynamoDB",         len(ddb_rows),  ddb_gib,  round(ddb_gib/1024,6),   pct_enc(ddb_rows,"Encrypted"),  inv_regions(ddb_rows)),
        ("Redshift",         len(rs_rows),   0,        0,                        pct_enc(rs_rows,"Encrypted"),   inv_regions(rs_rows)),
        ("EKS Clusters",     len(eks_rows),  0,        0,                        "—",                            inv_regions(eks_rows)),
        ("ECS Clusters",     len(ecs_rows),  0,        0,                        "—",                            inv_regions(ecs_rows)),
        ("Lambda",           len(lam_rows),  0,        0,                        "—",                            inv_regions(lam_rows)),
        ("WorkSpaces",       len(ws2_rows),  ws2_gib,  round(ws2_gib/1024,4),   "—",                            inv_regions(ws2_rows)),
        ("DocumentDB",       len(doc_rows),  0,        0,                        pct_enc(doc_rows,"Encrypted"),  inv_regions(doc_rows)),
        ("ElastiCache",      len(ec_rows),   0,        0,                        "—",                            inv_regions(ec_rows)),
    ]

    for i, row_vals in enumerate(inventory):
        bg = C_ALT_ROW if i % 2 == 0 else None
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=cur_l, column=ci, value=val)
            if bg:
                c.fill = hex_fill(bg)
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                    vertical="center", indent=1 if ci == 1 else 0)
        ws.row_dimensions[cur_l].height = 16
        cur_l += 1

    # Totals row
    for ci, val in enumerate([
        "TOTAL", total_res, total_gib, total_tib, "—", "—"
    ], 1):
        c = ws.cell(row=cur_l, column=ci, value=val)
        c.fill = hex_fill(C_SUMMARY_FILL)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center", indent=1 if ci == 1 else 0)
    ws.row_dimensions[cur_l].height = 18
    cur_l += 2  # spacer

    # ── LEFT: EC2 State Breakdown ─────────────────────────────────────────────
    cur_l = section_header(cur_l, 1, 6, "  EC2 Instance State Breakdown")
    state_map = Counter(r.get("State", "unknown") for r in ec2_rows)
    type_map  = Counter(r.get("Instance Type", "unknown") for r in ec2_rows)
    os_map    = Counter(r.get("OS / Platform", "Linux/Unix") for r in ec2_rows)

    for label, counter, total in [
        ("State",        state_map, len(ec2_rows)),
        ("Top Types",    type_map,  len(ec2_rows)),
        ("OS",           os_map,    len(ec2_rows)),
    ]:
        for i, (k, v) in enumerate(counter.most_common(5)):
            bg = C_ALT_ROW if i % 2 == 0 else None
            pct = f"{round(v/total*100)}%" if total else "—"
            bar = "█" * min(20, round(v / max(total, 1) * 20))
            for ci, val in enumerate([label if i == 0 else "", k, v, pct, bar, ""], 1):
                c = ws.cell(row=cur_l, column=ci, value=val)
                if bg: c.fill = hex_fill(bg)
                c.font = Font(size=9,
                              bold=(ci == 1 and i == 0),
                              color="2471A3" if ci == 5 else "000000")
                c.alignment = Alignment(vertical="center", indent=1 if ci <= 2 else 0,
                                        horizontal="left" if ci <= 2 else "center")
            ws.row_dimensions[cur_l].height = 15
            cur_l += 1
        cur_l += 1  # gap between groups

    # ── RIGHT: Risk & Findings ────────────────────────────────────────────────
    cur_r = section_header(cur_r, 7, 12, "  Risk & Findings")

    risks = [
        ("CRITICAL", "Unencrypted EBS Volumes",         len(unenc_ebs),
         f"{len(unenc_ebs)} volume(s) at risk of data exposure"),
        ("CRITICAL", "Public S3 Buckets",               len(pub_s3),
         f"{len(pub_s3)} bucket(s) not fully blocking public access"),
        ("CRITICAL", "Publicly Accessible RDS",         len(pub_rds),
         f"{len(pub_rds)} DB instance(s) exposed to internet"),
        ("HIGH",     "EC2 Without Backup Tag",          len(no_backup_ec2),
         f"{len(no_backup_ec2)} instance(s) missing backup=true tag"),
        ("HIGH",     "RDS Without Automated Backup",    len(rds_no_bk),
         f"{len(rds_no_bk)} DB(s) with 0-day retention — no recovery point"),
        ("HIGH",     "DynamoDB Without PITR",           len(ddb_no_pitr),
         f"{len(ddb_no_pitr)} table(s) cannot recover to point-in-time"),
        ("MEDIUM",   "S3 Buckets Without Versioning",   len(no_vers_s3),
         f"{len(no_vers_s3)} bucket(s) — accidental delete is unrecoverable"),
        ("MEDIUM",   "Unattached EBS Volumes",          len(unatt_ebs),
         f"{len(unatt_ebs)} volume(s) not in use — cost waste + backup gap"),
    ]

    risk_hdr = ["Severity", "Finding", "Count", "Detail"]
    for ci, h in enumerate(risk_hdr, 7):
        c = ws.cell(row=cur_r, column=ci, value=h)
        c.fill = hex_fill(C_SUBHDR_FILL)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[cur_r].height = 20
    cur_r += 1

    SEV_COLOR = {"CRITICAL": C_CRITICAL, "HIGH": C_WARN, "MEDIUM": "FEF9E7"}
    SEV_FONT  = {"CRITICAL": "922B21",   "HIGH": "7D6608", "MEDIUM": "4A235A"}

    for i, (sev, finding, cnt, detail) in enumerate(risks):
        bg = SEV_COLOR.get(sev, C_ALT_ROW)
        row_bg = C_ALT_ROW if i % 2 == 0 else None
        for ci, val in zip(range(7, 13), [sev, finding, cnt, detail, "", ""]):
            c = ws.cell(row=cur_r, column=ci, value=val)
            if ci == 7:
                c.fill = hex_fill(bg)
                c.font = Font(bold=True, size=9, color=SEV_FONT.get(sev, "000000"))
            else:
                if row_bg: c.fill = hex_fill(row_bg)
                c.font = Font(size=9,
                              bold=(cnt > 0 and ci == 9),
                              color="C0392B" if (cnt > 0 and ci == 9) else "000000")
            c.alignment = Alignment(horizontal="center" if ci in (7, 9) else "left",
                                    vertical="center", wrap_text=(ci == 10))
        ws.row_dimensions[cur_r].height = 18
        cur_r += 1

    cur_r += 1  # spacer

    # ── RIGHT: Backup Infrastructure ─────────────────────────────────────────
    cur_r = section_header(cur_r, 7, 12, "  AWS Backup Infrastructure")

    bk_data = [
        ("Backup Vaults",          len(bkv_rows),        ""),
        ("  — Immutable (Locked)", locked_vaults,         "WORM protection enabled"),
        ("Total Recovery Points",  total_rp,              "across all vaults"),
        ("Backup Plans",           len(bkp_rows),         ""),
        ("Cross-Region Copy Rules",
         sum(1 for r in bkp_rows if r.get("Copy To Region")), "DR copy configured"),
    ]

    for i, (label, val, note) in enumerate(bk_data):
        bg = C_ALT_ROW if i % 2 == 0 else None
        for ci, v in zip(range(7, 13), [label, val, note, "", "", ""]):
            c = ws.cell(row=cur_r, column=ci, value=v)
            if bg: c.fill = hex_fill(bg)
            c.font = Font(size=9, bold=(ci == 7),
                          color="1E8449" if (ci == 8 and isinstance(val, int) and val > 0) else "000000")
            c.alignment = Alignment(horizontal="left" if ci in (7, 9) else "center",
                                    vertical="center", indent=1 if ci == 7 else 0)
        ws.row_dimensions[cur_r].height = 16
        cur_r += 1

    cur_r += 1

    # ── RIGHT: Region Distribution ────────────────────────────────────────────
    cur_r = section_header(cur_r, 7, 12, "  Resource Distribution by Region")

    for ci, h in enumerate(["Region", "Resources", "Share", "", "", ""], 7):
        c = ws.cell(row=cur_r, column=ci, value=h)
        if h:
            c.fill = hex_fill(C_SUBHDR_FILL)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[cur_r].height = 20
    cur_r += 1

    total_placed = sum(region_counts.values()) or 1
    for i, (reg, cnt) in enumerate(region_counts.most_common()):
        bg = C_ALT_ROW if i % 2 == 0 else None
        pct_val = round(cnt / total_placed * 100)
        bar = "█" * min(18, round(cnt / total_placed * 18))
        for ci, val in zip(range(7, 13), [reg, cnt, f"{pct_val}%", bar, "", ""]):
            c = ws.cell(row=cur_r, column=ci, value=val)
            if bg: c.fill = hex_fill(bg)
            c.font = Font(size=9,
                          color="2471A3" if ci == 10 else "000000")
            c.alignment = Alignment(horizontal="left" if ci in (7, 10) else "center",
                                    vertical="center")
        ws.row_dimensions[cur_r].height = 15
        cur_r += 1

    # ── RIGHT: Storage by Service ─────────────────────────────────────────────
    cur_r += 1
    cur_r = section_header(cur_r, 7, 12, "  Storage by Service (GiB)")

    storage_breakdown = [
        ("EBS Volumes",   ebs_gib),
        ("S3 Buckets",    s3_gib),
        ("RDS / Aurora",  rds_gib),
        ("EC2 (root)",    ec2_gib),
        ("EFS",           efs_gib),
        ("FSx",           fsx_gib),
        ("DynamoDB",      ddb_gib),
        ("WorkSpaces",    ws2_gib),
    ]
    storage_breakdown.sort(key=lambda x: x[1], reverse=True)
    max_stor = max((v for _, v in storage_breakdown), default=1) or 1

    for i, (svc, gib_val) in enumerate(storage_breakdown):
        bg = C_ALT_ROW if i % 2 == 0 else None
        pct_val = round(gib_val / max(total_gib, 0.001) * 100)
        bar = "█" * min(16, round(gib_val / max_stor * 16))
        for ci, val in zip(range(7, 13),
                           [svc, gib_val, f"{pct_val}%", bar, "", ""]):
            c = ws.cell(row=cur_r, column=ci, value=val)
            if bg: c.fill = hex_fill(bg)
            c.font = Font(size=9, color="2471A3" if ci == 10 else "000000")
            c.alignment = Alignment(horizontal="left" if ci in (7, 10) else "center",
                                    vertical="center")
        ws.row_dimensions[cur_r].height = 15
        cur_r += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {
        1: 18, 2: 8, 3: 14, 4: 14, 5: 12, 6: 22,
        7: 22, 8: 10, 9: 22, 10: 20, 11: 6, 12: 6,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A9"


# ─── Main orchestrator ────────────────────────────────────────────────────────

def collect_region(session, region, account_id):
    result = {}
    collectors = [
        ("EC2 Instances",   lambda: collect_ec2(session, region)),
        ("EBS Volumes",     lambda: collect_ebs_volumes(session, region)),
        ("EBS Snapshots",   lambda: collect_ebs_snapshots(session, region, account_id)),
        ("RDS & Aurora",    lambda: collect_rds(session, region)),
        ("EFS",             lambda: collect_efs(session, region)),
        ("FSx",             lambda: collect_fsx(session, region)),
        ("DynamoDB",        lambda: collect_dynamodb(session, region)),
        ("Redshift",        lambda: collect_redshift(session, region)),
        ("EKS",             lambda: collect_eks(session, region)),
        ("ECS",             lambda: collect_ecs(session, region)),
        ("Lambda",          lambda: collect_lambda(session, region)),
        ("WorkSpaces",      lambda: collect_workspaces(session, region)),
        ("DocumentDB",      lambda: collect_documentdb(session, region)),
        ("ElastiCache",     lambda: collect_elasticache(session, region)),
        ("AWS Backup Vaults",  lambda: collect_aws_backup(session, region)),
        ("AWS Backup Plans",   lambda: collect_backup_plans(session, region)),
    ]
    for name, fn in collectors:
        try:
            result[name] = fn()
        except Exception as e:
            log.warning("Collector '%s' in %s failed: %s", name, region, e)
            result[name] = []
    return result


def merge_results(all_results):
    merged = defaultdict(list)
    for region_result in all_results:
        for key, rows in region_result.items():
            merged[key].extend(rows)
    return dict(merged)


SHEET_COLS = {
    "EC2 Instances": [
        "Region", "Name", "Instance ID", "State", "Instance Type", "OS / Platform",
        "AZ", "Environment", "Owner", "Volume Count", "Root Disk (GiB)",
        "Data Disks (GiB)", "Total Storage (GiB)", "VPC ID", "Subnet ID",
        "AMI ID", "Launch Time", "Tag:Backup", "Notes",
    ],
    "EBS Volumes": [
        "Region", "Volume ID", "Name", "State", "Type", "Size (GiB)", "IOPS",
        "Throughput", "Encrypted", "Multi-Attach", "AZ", "Attached To",
        "Snapshot ID", "Created",
    ],
    "EBS Snapshots": [
        "Region", "Snapshot ID", "Name", "Volume ID", "State", "Size (GiB)",
        "Encrypted", "Description", "Start Time",
    ],
    "RDS & Aurora": [
        "Region", "DB Identifier", "Engine", "Engine Version", "Instance Class",
        "Status", "Multi-AZ", "Storage Type", "Allocated Storage (GiB)",
        "Max Allocated (GiB)", "Encrypted", "Backup Retention (days)",
        "Automated Backups", "DB Cluster ID", "VPC", "AZ", "License Model",
        "Public", "Created", "Notes",
    ],
    "S3 Buckets": [
        "Region", "Bucket Name", "Created", "Size (MiB)", "Size (GiB)", "Size (TiB)",
        "Object Count", "Versioning", "Replication", "Lifecycle Rules",
        "Encryption", "Public Access", "Notes",
    ],
    "EFS": [
        "Region", "File System ID", "Name", "State", "Performance Mode",
        "Throughput Mode", "Provisioned Throughput (MiBps)", "Encrypted",
        "Size (GiB)", "IA Size (GiB)", "Standard Size (GiB)", "Created",
    ],
    "FSx": [
        "Region", "File System ID", "Name", "Type", "State", "Storage Type",
        "Capacity (GiB)", "Encrypted", "VPC", "AZs", "Configuration", "Created",
    ],
    "DynamoDB": [
        "Region", "Table Name", "Status", "Billing Mode", "Size (GiB)",
        "Item Count", "RCU", "WCU", "PITR Enabled", "Global Tables",
        "Streams", "Encrypted", "Created",
    ],
    "Redshift": [
        "Region", "Cluster ID", "Status", "Node Type", "Nodes", "DB Name",
        "Encrypted", "Backup Retention (days)", "Automated Backups", "Public",
        "VPC", "AZ", "Serverless", "Created",
    ],
    "EKS": [
        "Region", "Cluster Name", "Status", "K8s Version", "Platform Version",
        "Node Groups", "Total Nodes", "Node Details", "VPC",
        "Private Endpoint", "Logging", "Created",
    ],
    "ECS": [
        "Region", "Cluster Name", "Status", "Active Services", "Running Tasks",
        "Pending Tasks", "Registered Instances", "Capacity Providers", "Notes",
    ],
    "Lambda": [
        "Region", "Function Name", "Runtime", "Memory (MB)", "Timeout (sec)",
        "Package Size (MB)", "Architecture", "Last Modified", "Handler",
        "Description",
    ],
    "WorkSpaces": [
        "Region", "Workspace ID", "User", "State", "Bundle ID", "Directory ID",
        "Running Mode", "Root Volume (GiB)", "User Volume (GiB)", "Compute Type",
        "Protocol",
    ],
    "DocumentDB": [
        "Region", "Cluster ID", "Engine", "Engine Version", "Status", "Members",
        "Storage (GiB)", "Encrypted", "Backup Retention (days)", "Multi-AZ",
        "VPC", "Created",
    ],
    "ElastiCache": [
        "Region", "Cluster ID", "Engine", "Engine Version", "Node Type", "Status",
        "Nodes", "AZ", "Replication Group", "Encrypted at Rest",
        "Encrypted in Transit", "Backup Retention (days)", "Created",
    ],
    "AWS Backup Vaults": [
        "Region", "Vault Name", "Recovery Points", "Encrypted", "Locked",
        "Min Retention (days)", "Max Retention (days)", "Created",
    ],
    "AWS Backup Plans": [
        "Region", "Plan Name", "Rule Name", "Target Vault", "Schedule",
        "Start Window (min)", "Completion Window (min)", "Delete After (days)",
        "Cold After (days)", "Copy To Region", "Created",
    ],
}


def ec2_color_row(cell, col, row):
    if col == "State" and row.get("State") == "stopped":
        cell.fill = hex_fill(C_WARN)
    if col == "Total Storage (GiB)":
        val = row.get("Total Storage (GiB)", 0) or 0
        if val > 5000:
            cell.fill = hex_fill(C_WARN)
        elif val > 10000:
            cell.fill = hex_fill(C_CRITICAL)


def rds_color_row(cell, col, row):
    if col == "Automated Backups" and row.get("Automated Backups") == "No":
        cell.fill = hex_fill(C_CRITICAL)
    if col == "Backup Retention (days)" and (row.get("Backup Retention (days)") or 0) == 0:
        cell.fill = hex_fill(C_CRITICAL)
    if col == "Multi-AZ" and not row.get("Multi-AZ"):
        cell.fill = hex_fill(C_WARN)


def s3_color_row(cell, col, row):
    if col == "Public Access" and row.get("Public Access") not in ("Blocked", ""):
        cell.fill = hex_fill(C_CRITICAL)
    if col == "Versioning" and row.get("Versioning") in ("Disabled", "Suspended", ""):
        cell.fill = hex_fill(C_WARN)


COLOR_FUNCS = {
    "EC2 Instances": ec2_color_row,
    "RDS & Aurora":  rds_color_row,
    "S3 Buckets":    s3_color_row,
}


def build_workbook(data, account_id, regions, assessed_at, output_path):
    wb = openpyxl.Workbook()
    build_summary_sheet(wb, data, account_id, regions, assessed_at)

    for sheet_name, columns in SHEET_COLS.items():
        rows = data.get(sheet_name, [])
        color_fn = COLOR_FUNCS.get(sheet_name)
        add_sheet(
            wb,
            name=sheet_name,
            rows=rows,
            columns=columns,
            color_row=color_fn,
            title=f"{sheet_name}  ({len(rows)} resources)",
        )

    wb.save(output_path)
    print(f"\n✓ Assessment saved: {output_path}")


# ─── CLI ──────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description=(
            "AWS Environment Assessment Tool — read-only inventory scanner that "
            "produces a multi-sheet Excel workbook covering every major AWS workload type."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--regions", nargs="+", default=None,
        help="List of AWS regions to scan (default: current region only)",
    )
    p.add_argument(
        "--all-regions", action="store_true",
        help="Scan all enabled regions in the account",
    )
    p.add_argument(
        "--profile", default=None,
        help="AWS CLI profile name (default: default profile)",
    )
    p.add_argument(
        "--output", default=None,
        help="Output xlsx filename (default: aws_assessment_<account>_<date>.xlsx)",
    )
    p.add_argument(
        "--workers", type=int, default=4,
        help="Parallel region workers (default: 4)",
    )
    p.add_argument(
        "--skip-snapshots", action="store_true",
        help="Skip EBS snapshot enumeration (can be slow on large accounts)",
    )
    p.add_argument(
        "--verbose", action="store_true",
        help="Enable verbose logging",
    )
    return p.parse_args()


def main():
    args = parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.INFO)

    # Session
    session = boto3.Session(profile_name=args.profile)
    account_id = get_account_id(session)
    print(f"Account: {account_id}")

    # Regions
    if args.all_regions:
        regions = get_all_regions(session)
    elif args.regions:
        regions = args.regions
    else:
        regions = [session.region_name or "us-east-1"]

    print(f"Regions: {', '.join(regions)}")
    assessed_at = now_str()

    # S3 is global — collect once
    print("\nCollecting S3 (global)...")
    s3_data = collect_s3(session)

    # Patch: skip snapshots if requested
    if args.skip_snapshots:
        SHEET_COLS.pop("EBS Snapshots", None)

    # Parallel region collection
    print(f"\nCollecting regional resources across {len(regions)} region(s)...")
    all_results = []

    def collect_one(region):
        print(f"  → {region}")
        return collect_region(session, region, account_id)

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(collect_one, r): r for r in regions}
        for future in as_completed(futures):
            region = futures[future]
            try:
                all_results.append(future.result())
            except Exception as e:
                print(f"  ✗ {region}: {e}")

    # Merge
    data = merge_results(all_results)
    data["S3 Buckets"] = s3_data

    # Output path — strip any directory components to prevent path traversal
    import os as _os
    date_str = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%d")
    raw_output = args.output or f"aws_assessment_{account_id}_{date_str}.xlsx"
    output = _os.path.basename(raw_output) if args.output else raw_output
    if not output.endswith(".xlsx"):
        output += ".xlsx"

    # Build workbook
    print("\nBuilding Excel workbook...")
    build_workbook(data, account_id, regions, assessed_at, output)

    # Print quick totals
    print("\n─── Quick Totals ────────────────────────────────────────")
    for key in SHEET_COLS:
        count = len(data.get(key, []))
        if count:
            print(f"  {key:<25} {count:>6} resources")
    print("─────────────────────────────────────────────────────────")


if __name__ == "__main__":
    main()
