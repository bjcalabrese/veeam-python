#!/usr/bin/env python3
"""
Azure Environment Assessment Tool

Read-only inventory tool that scans an Azure subscription and produces a single
multi-sheet Excel workbook covering every major workload type — Virtual Machines,
Managed Disks, Disk Snapshots, Azure SQL, Cosmos DB, Storage Accounts, Azure
NetApp Files, Synapse Analytics, AKS, Container Instances, Azure Functions,
Azure Virtual Desktop, Azure Cache for Redis, and Azure Backup.

DISCLAIMER
----------
This is a community sample script provided without support guarantees.
It is not an official product and is not covered by any support agreement.
Use at your own risk. Review the code before running it in any environment.

Requirements:
    pip install -r requirements.txt

Usage:
    python azure_assessment.py
    python azure_assessment.py --subscription 00000000-0000-0000-0000-000000000000
    python azure_assessment.py --all-subscriptions --output my_assessment.xlsx
"""

import os as _os
import sys
import argparse
import datetime
import logging
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

# ─── AZURE SDK ────────────────────────────────────────────────────────────────
try:
    from azure.identity import DefaultAzureCredential
    from azure.core.exceptions import HttpResponseError
except ImportError as exc:
    print(f"Missing azure-identity: {exc}\nRun: pip install -r requirements.txt")
    sys.exit(1)

def _try_import(module, cls):
    """Safely import an Azure management client class; return None if unavailable."""
    try:
        import importlib
        return getattr(importlib.import_module(module), cls)
    except (ImportError, AttributeError):
        return None

_SubscriptionClient   = _try_import("azure.mgmt.resource",              "SubscriptionClient")
_ComputeClient        = _try_import("azure.mgmt.compute",               "ComputeManagementClient")
_StorageClient        = _try_import("azure.mgmt.storage",               "StorageManagementClient")
_SqlClient            = _try_import("azure.mgmt.sql",                   "SqlManagementClient")
_AKSClient            = _try_import("azure.mgmt.containerservice",      "ContainerServiceClient")
_WebClient            = _try_import("azure.mgmt.web",                   "WebSiteManagementClient")
_CosmosClient         = _try_import("azure.mgmt.cosmosdb",              "CosmosDBManagementClient")
_RecoveryClient       = _try_import("azure.mgmt.recoveryservices",      "RecoveryServicesClient")
_BackupClient         = _try_import("azure.mgmt.recoveryservicesbackup","RecoveryServicesBackupClient")
_RedisClient          = _try_import("azure.mgmt.redis",                 "RedisManagementClient")
_NetAppClient         = _try_import("azure.mgmt.netapp",                "NetAppManagementClient")
_AVDClient            = _try_import("azure.mgmt.desktopvirtualization", "DesktopVirtualizationMgmtClient")
_SynapseClient        = _try_import("azure.mgmt.synapse",               "SynapseManagementClient")
_ACIClient            = _try_import("azure.mgmt.containerinstance",     "ContainerInstanceManagementClient")
_MonitorClient        = _try_import("azure.mgmt.monitor",               "MonitorManagementClient")
_SqlVMClient          = _try_import("azure.mgmt.sqlvirtualmachine",     "SqlVirtualMachineManagementClient")
_CostClient           = _try_import("azure.mgmt.costmanagement",        "CostManagementClient")

# Critical imports — fail if missing
for _name, _obj in [("azure-mgmt-resource", _SubscriptionClient),
                    ("azure-mgmt-compute",  _ComputeClient),
                    ("azure-mgmt-storage",  _StorageClient)]:
    if _obj is None:
        print(f"ERROR: {_name} not installed.  Run: pip install -r requirements.txt")
        sys.exit(1)

# ─── EXCEL ────────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing openpyxl.  Run: pip install -r requirements.txt")
    sys.exit(1)

try:
    from tqdm import tqdm
except ImportError:
    def tqdm(it, **_):
        return it

# ─── LOGGING ──────────────────────────────────────────────────────────────────
log = logging.getLogger("azure_assessment")

# ─── STYLE CONSTANTS ──────────────────────────────────────────────────────────
RED_FILL    = PatternFill("solid", fgColor="FFB3B3")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2B3")
GREEN_FILL  = PatternFill("solid", fgColor="B3FFB3")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
TITLE_FILL  = PatternFill("solid", fgColor="0070C0")
KPI_FILL    = PatternFill("solid", fgColor="2E75B6")
CRIT_FILL   = PatternFill("solid", fgColor="C00000")
HIGH_FILL   = PatternFill("solid", fgColor="FF6600")
MED_FILL    = PatternFill("solid", fgColor="FFC000")

HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
TITLE_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=14)
KPI_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=20)
KPI_LBL     = Font(bold=False, color="FFFFFF", name="Calibri", size=9)
BOLD        = Font(bold=True, name="Calibri", size=10)
NORMAL      = Font(name="Calibri", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def gib(b):
    return round(b / 1_073_741_824, 4) if b else 0.0

def tib(b):
    return round(b / 1_099_511_627_776, 6) if b else 0.0

def mib(b):
    return round(b / 1_048_576, 4) if b else 0.0

def tag(tags, *keys):
    """Extract first matching tag key (case-insensitive) from an Azure tags dict."""
    if not tags:
        return ""
    tags_lower = {k.lower(): v for k, v in tags.items()}
    for k in keys:
        v = tags_lower.get(k.lower(), "")
        if v:
            return v
    return ""

def rg_from_id(resource_id):
    """Extract resource group name from an Azure resource ID."""
    try:
        parts = resource_id.split("/")
        idx = [p.lower() for p in parts].index("resourcegroups")
        return parts[idx + 1]
    except (ValueError, IndexError):
        return ""

def safe_list(iterable):
    """Consume an Azure SDK iterable, swallowing errors."""
    results = []
    try:
        for item in iterable:
            results.append(item)
    except Exception:
        pass
    return results

def storage_metric(monitor, resource_id, metric_name, namespace):
    """Fetch the latest average value of an Azure Monitor metric in bytes."""
    if monitor is None:
        return None
    try:
        end   = datetime.datetime.now(datetime.timezone.utc)
        start = end - datetime.timedelta(days=2)
        ts    = f"{start.strftime('%Y-%m-%dT%H:%M:%SZ')}/{end.strftime('%Y-%m-%dT%H:%M:%SZ')}"
        res   = monitor.metrics.list(
            resource_id,
            timespan=ts,
            interval="PT1H",
            metricnames=metric_name,
            aggregation="Average",
            metricnamespace=namespace,
        )
        for metric in res.value:
            for series in metric.timeseries:
                for dp in reversed(series.data):
                    if dp.average is not None:
                        return dp.average
    except Exception:
        pass
    return None

def blob_capacity_by_tier(monitor, resource_id):
    """Fetch BlobCapacity broken down by storage tier (Hot/Cool/Cold/Archive).
    Returns dict {tier_name: bytes}.  One API call returns all tiers at once.
    """
    result = {}
    if monitor is None:
        return result
    try:
        end   = datetime.datetime.now(datetime.timezone.utc)
        start = end - datetime.timedelta(days=2)
        ts    = f"{start.strftime('%Y-%m-%dT%H:%M:%SZ')}/{end.strftime('%Y-%m-%dT%H:%M:%SZ')}"
        res   = monitor.metrics.list(
            resource_id,
            timespan=ts,
            interval="PT1H",
            metricnames="BlobCapacity",
            aggregation="Average",
            metricnamespace="microsoft.storage/storageaccounts/blobservices",
            filter="Tier ne 'Unknown'",        # returns a timeseries per tier
        )
        for metric in res.value:
            for series in metric.timeseries:
                tier_name = ""
                for mv in (series.metadatavalues or []):
                    if mv.name and mv.name.value and mv.name.value.lower() == "tier":
                        tier_name = (mv.value or "").capitalize()
                        break
                if not tier_name:
                    continue
                for dp in reversed(series.data):
                    if dp.average is not None:
                        result[tier_name] = dp.average
                        break
    except Exception:
        pass
    return result

# ─── COLLECTORS ───────────────────────────────────────────────────────────────

def collect_vms(credential, sub_id, sub_name, verbose=False):
    """Collect all Virtual Machines in the subscription."""
    rows = []
    if _ComputeClient is None:
        return rows
    try:
        compute = _ComputeClient(credential, sub_id)
        vms = safe_list(compute.virtual_machines.list_all())

        def _power_state(vm):
            try:
                rg = rg_from_id(vm.id)
                iv = compute.virtual_machines.instance_view(rg, vm.name)
                for s in iv.statuses:
                    if s.code and s.code.startswith("PowerState/"):
                        return s.display_status or s.code.split("/")[-1]
            except Exception:
                pass
            return "Unknown"

        with ThreadPoolExecutor(max_workers=10) as exe:
            states = dict(zip([v.name for v in vms], exe.map(_power_state, vms)))

        for vm in vms:
            rg       = rg_from_id(vm.id)
            location = vm.location or ""
            size     = (vm.hardware_profile.vm_size if vm.hardware_profile else "") or ""
            os_type  = ""
            os_disk  = 0
            data_cnt = 0
            data_gib = 0
            zones    = ",".join(vm.zones) if vm.zones else ""

            if vm.storage_profile:
                od = vm.storage_profile.os_disk
                if od:
                    os_type = str(od.os_type or "")
                    os_disk = od.disk_size_gb or 0
                dds = vm.storage_profile.data_disks or []
                data_cnt = len(dds)
                data_gib = sum(d.disk_size_gb or 0 for d in dds)

            tags_ = vm.tags or {}
            rows.append({
                "Subscription": sub_name,
                "Name":               vm.name,
                "Resource Group":     rg,
                "Location":           location,
                "VM Size":            size,
                "OS Type":            os_type,
                "Power State":        states.get(vm.name, "Unknown"),
                "OS Disk (GiB)":      os_disk,
                "Data Disks":         data_cnt,
                "Data Disk (GiB)":    data_gib,
                "Total Storage (GiB)":os_disk + data_gib,
                "Zones":              zones,
                "Environment":        tag(tags_, "Environment", "env", "Env"),
                "Owner":              tag(tags_, "Owner", "owner"),
                "Backup Tag":         tag(tags_, "Backup", "backup", "BackupPolicy"),
                "MSSQL-INSTALLED":    "",   # filled in post-collection
                "Backup Policy":      "",   # filled in post-collection
                "Backup Protected":   "",   # filled in post-collection
            })
        if verbose:
            log.info("VMs %s: %d found", sub_name, len(rows))
    except Exception as exc:
        log.warning("VMs %s: %s", sub_name, exc)
    return rows


def collect_disks(credential, sub_id, sub_name, verbose=False):
    """Collect all Managed Disks in the subscription."""
    rows = []
    if _ComputeClient is None:
        return rows
    try:
        compute = _ComputeClient(credential, sub_id)
        for d in safe_list(compute.disks.list()):
            rg        = rg_from_id(d.id)
            sku       = d.sku.name if d.sku else ""
            size_gib  = d.disk_size_gb or 0
            iops      = d.disk_iops_read_write or 0
            mbps      = d.disk_m_bps_read_write or 0
            state     = str(d.disk_state or "")
            attached  = ""
            if d.managed_by:
                attached = d.managed_by.split("/")[-1]
            enc_type  = ""
            if d.encryption:
                enc_type = str(d.encryption.type or "")
            tags_     = d.tags or {}
            # AKS Persistent Volume detection via Kubernetes tags on the disk
            pvc_name  = tag(tags_, "kubernetes.io-created-for-pvc-name",
                                    "kubernetes.io/created-for/pvc/name")
            pvc_ns    = tag(tags_, "kubernetes.io-created-for-pvc-namespace",
                                    "kubernetes.io/created-for/pvc/namespace")
            pv_name   = tag(tags_, "kubernetes.io-created-for-pv-name",
                                    "kubernetes.io/created-for/pv/name")
            is_pv     = "Yes" if (pvc_name or pv_name) else ""
            rows.append({
                "Subscription":   sub_name,
                "Name":           d.name,
                "Resource Group": rg,
                "Location":       d.location or "",
                "SKU":            sku,
                "Size (GiB)":     size_gib,
                "IOPS":           iops,
                "Throughput (MB/s)": mbps,
                "Encryption Type":enc_type,
                "Disk State":     state,
                "Attached To":    attached,
                "AKS PV":         is_pv,
                "PVC Name":       pvc_name,
                "PVC Namespace":  pvc_ns,
                "Environment":    tag(tags_, "Environment", "env"),
            })
        if verbose:
            log.info("Disks %s: %d found", sub_name, len(rows))
    except Exception as exc:
        log.warning("Disks %s: %s", sub_name, exc)
    return rows


def collect_snapshots(credential, sub_id, sub_name, verbose=False):
    """Collect all Disk Snapshots in the subscription."""
    rows = []
    if _ComputeClient is None:
        return rows
    try:
        compute = _ComputeClient(credential, sub_id)
        now     = datetime.datetime.now(datetime.timezone.utc)
        for s in safe_list(compute.snapshots.list()):
            rg       = rg_from_id(s.id)
            size_gib = s.disk_size_gb or 0
            enc      = ""
            if s.encryption:
                enc = str(s.encryption.type or "")
            created  = s.time_created
            age_days = (now - created.replace(tzinfo=datetime.timezone.utc)).days if created else ""
            rows.append({
                "Subscription":   sub_name,
                "Name":           s.name,
                "Resource Group": rg,
                "Location":       s.location or "",
                "Source Disk":    (s.creation_data.source_resource_id or "").split("/")[-1] if s.creation_data else "",
                "Size (GiB)":     size_gib,
                "Encryption":     enc,
                "Created":        str(created.date()) if created else "",
                "Age (days)":     age_days,
            })
        if verbose:
            log.info("Snapshots %s: %d found", sub_name, len(rows))
    except Exception as exc:
        log.warning("Snapshots %s: %s", sub_name, exc)
    return rows


def collect_sql(credential, sub_id, sub_name, verbose=False):
    """Collect Azure SQL Databases and Managed Instances."""
    rows = []
    if _SqlClient is None:
        return rows
    try:
        sql     = _SqlClient(credential, sub_id)
        monitor = _MonitorClient(credential, sub_id) if _MonitorClient else None

        def _sql_metrics(resource_id):
            """Return (allocated_gib, used_gib) from Monitor; None if unavailable."""
            if not monitor:
                return None, None
            try:
                end   = datetime.datetime.now(datetime.timezone.utc)
                start = end - datetime.timedelta(days=1)
                ts    = f"{start.strftime('%Y-%m-%dT%H:%M:%SZ')}/{end.strftime('%Y-%m-%dT%H:%M:%SZ')}"
                res   = monitor.metrics.list(
                    resource_id, timespan=ts, interval="PT1H",
                    metricnames="allocated_data_storage,storage",
                    aggregation="Maximum",
                    metricnamespace="microsoft.sql/servers/databases",
                )
                alloc = used = None
                for m in res.value:
                    for series in m.timeseries:
                        for dp in reversed(series.data):
                            if dp.maximum is not None:
                                val = round(dp.maximum / 1_073_741_824, 4)
                                if "allocated" in m.name.value.lower():
                                    alloc = val
                                else:
                                    used = val
                                break
                return alloc, used
            except Exception:
                return None, None

        def _sql_retention(rg, server, db_name):
            """Return (pitr_days, ltr_weekly, ltr_monthly, ltr_yearly)."""
            pitr = ltr_w = ltr_m = ltr_y = ""
            try:
                p = sql.backup_short_term_retention_policies.get(rg, server, db_name)
                pitr = p.retention_days or ""
            except Exception:
                pass
            try:
                p = sql.backup_long_term_retention_policies.get(rg, server, db_name)
                ltr_w = p.weekly_retention  or ""
                ltr_m = p.monthly_retention or ""
                ltr_y = p.yearly_retention  or ""
            except Exception:
                pass
            return pitr, ltr_w, ltr_m, ltr_y

        # PaaS databases
        for srv in safe_list(sql.servers.list()):
            rg_s = rg_from_id(srv.id)
            for db in safe_list(sql.databases.list_by_server(rg_s, srv.name)):
                if db.name == "master":
                    continue
                sku_name   = db.sku.name if db.sku else ""
                tier       = db.sku.tier if db.sku else ""
                cap        = db.sku.capacity if db.sku else ""
                storage    = (db.max_size_bytes or 0) / 1_073_741_824
                retention  = getattr(db, "requested_backup_storage_redundancy", None) or getattr(db, "backup_storage_redundancy", None) or ""
                pool_name  = getattr(db, "elastic_pool_name", None) or ""
                alloc, used = _sql_metrics(db.id)
                pitr, ltr_w, ltr_m, ltr_y = _sql_retention(rg_s, srv.name, db.name)
                rows.append({
                    "Subscription":      sub_name,
                    "Type":              "Azure SQL Database",
                    "Server / Instance": srv.name,
                    "Database Name":     db.name,
                    "Resource Group":    rg_s,
                    "Location":          db.location or "",
                    "SKU":               sku_name,
                    "Tier":              tier,
                    "Capacity":          str(cap),
                    "Max Storage (GiB)": round(storage, 2),
                    "Allocated (GiB)":   alloc,
                    "Used (GiB)":        used,
                    "Elastic Pool":      pool_name,
                    "PITR (Days)":       pitr,
                    "LTR Weekly":        ltr_w,
                    "LTR Monthly":       ltr_m,
                    "LTR Yearly":        ltr_y,
                    "Backup Redundancy": str(retention),
                    "Public Access":     str(srv.public_network_access or ""),
                    "Encryption (TDE)":  "Enabled",
                    "Availability Zone": getattr(db, "availability_zone", None) or "",
                })

        # Managed Instances
        for mi in safe_list(sql.managed_instances.list()):
            rg_m   = rg_from_id(mi.id)
            storage= mi.storage_size_in_gb or 0
            rows.append({
                "Subscription":      sub_name,
                "Type":              "SQL Managed Instance",
                "Server / Instance": mi.name,
                "Database Name":     "",
                "Resource Group":    rg_m,
                "Location":          mi.location or "",
                "SKU":               mi.sku.name if mi.sku else "",
                "Tier":              mi.sku.tier if mi.sku else "",
                "Capacity":          str(mi.v_cores or ""),
                "Max Storage (GiB)": storage,
                "Allocated (GiB)":   None,
                "Used (GiB)":        None,
                "Elastic Pool":      "",
                "PITR (Days)":       getattr(mi, "backup_storage_redundancy", None) and "" or "",
                "LTR Weekly":        "",
                "LTR Monthly":       "",
                "LTR Yearly":        "",
                "Backup Redundancy": str(mi.storage_account_type or ""),
                "Public Access":     str(mi.public_data_endpoint_enabled or False),
                "Encryption (TDE)":  "Enabled",
                "Availability Zone": getattr(mi, "zone_redundant", None) and "Zone-redundant" or "",
                "License Type":      str(getattr(mi, "license_type", "") or ""),
                "Proxy Override":    str(getattr(mi, "proxy_override", "") or ""),
                "Min TLS":           str(getattr(mi, "minimal_tls_version", "") or ""),
                "Timezone":          getattr(mi, "timezone_id", "") or "",
            })

        if verbose:
            log.info("SQL %s: %d entries", sub_name, len(rows))
    except Exception as exc:
        log.warning("SQL %s: %s", sub_name, exc)
    return rows


def collect_sql_mi_databases(credential, sub_id, sub_name, verbose=False):
    """Collect individual databases on each SQL Managed Instance, including backup retention."""
    rows = []
    if _SqlClient is None:
        return rows
    try:
        sql = _SqlClient(credential, sub_id)
        for mi in safe_list(sql.managed_instances.list()):
            rg       = rg_from_id(mi.id)
            vcores   = mi.v_cores or ""
            storage  = mi.storage_size_in_gb or 0
            location = mi.location or ""
            license_type = str(getattr(mi, "license_type", "") or "")
            try:
                dbs = safe_list(sql.managed_databases.list_by_instance(rg, mi.name))
            except Exception:
                continue
            for db in dbs:
                pitr = ltr_w = ltr_m = ltr_y = ""
                try:
                    p = sql.managed_backup_short_term_retention_policies.get(rg, mi.name, db.name)
                    pitr = p.retention_days or ""
                except Exception:
                    pass
                try:
                    p = sql.managed_backup_long_term_retention_policies.get(rg, mi.name, db.name)
                    ltr_w = getattr(p, "weekly_retention",  None) or ""
                    ltr_m = getattr(p, "monthly_retention", None) or ""
                    ltr_y = getattr(p, "yearly_retention",  None) or ""
                except Exception:
                    pass
                rows.append({
                    "Subscription":           sub_name,
                    "Managed Instance":       mi.name,
                    "Database Name":          db.name,
                    "Resource Group":         rg,
                    "Location":               location,
                    "Instance vCores":        vcores,
                    "Instance Storage (GiB)": storage,
                    "License Type":           license_type,
                    "DB Status":              str(db.status or ""),
                    "Collation":              db.collation or "",
                    "Created":                str(db.creation_date.date() if db.creation_date else ""),
                    "Earliest Restore":       str(db.earliest_restore_point.date() if db.earliest_restore_point else ""),
                    "PITR (Days)":            pitr,
                    "LTR Weekly":             ltr_w,
                    "LTR Monthly":            ltr_m,
                    "LTR Yearly":             ltr_y,
                })
        if verbose:
            log.info("SQL MI Databases %s: %d databases", sub_name, len(rows))
    except Exception as exc:
        log.warning("SQL MI Databases %s: %s", sub_name, exc)
    return rows


def collect_sql_elastic_pools(credential, sub_id, sub_name, verbose=False):
    """Collect SQL Elastic Pools with pool-level sizing and per-database assignments."""
    rows = []
    if _SqlClient is None:
        return rows
    try:
        sql     = _SqlClient(credential, sub_id)
        monitor = _MonitorClient(credential, sub_id) if _MonitorClient else None

        def _pool_metrics(resource_id):
            """Return (allocated_gib, used_gib) for a pool from Monitor."""
            if not monitor:
                return None, None
            try:
                end   = datetime.datetime.now(datetime.timezone.utc)
                start = end - datetime.timedelta(days=1)
                ts    = f"{start.strftime('%Y-%m-%dT%H:%M:%SZ')}/{end.strftime('%Y-%m-%dT%H:%M:%SZ')}"
                res   = monitor.metrics.list(
                    resource_id, timespan=ts, interval="PT1H",
                    metricnames="allocated_data_storage,storage_used",
                    aggregation="Maximum",
                    metricnamespace="microsoft.sql/servers/elasticpools",
                )
                alloc = used = None
                for m in res.value:
                    for series in m.timeseries:
                        for dp in reversed(series.data):
                            if dp.maximum is not None:
                                val = round(dp.maximum / 1_073_741_824, 4)
                                if "allocated" in m.name.value.lower():
                                    alloc = val
                                else:
                                    used = val
                                break
                return alloc, used
            except Exception:
                return None, None

        for srv in safe_list(sql.servers.list()):
            rg_s = rg_from_id(srv.id)
            for pool in safe_list(sql.elastic_pools.list_by_server(rg_s, srv.name)):
                sku_name   = pool.sku.name     if pool.sku else ""
                tier       = pool.sku.tier     if pool.sku else ""
                capacity   = pool.sku.capacity if pool.sku else ""
                max_gib    = (pool.max_size_bytes or 0) / 1_073_741_824
                alloc, used = _pool_metrics(pool.id)
                # Count databases in this pool
                try:
                    db_count = sum(1 for _ in sql.databases.list_by_elastic_pool(rg_s, srv.name, pool.name))
                except Exception:
                    db_count = ""
                rows.append({
                    "Subscription":      sub_name,
                    "Server":            srv.name,
                    "Pool Name":         pool.name,
                    "Resource Group":    rg_s,
                    "Location":          pool.location or "",
                    "SKU":               sku_name,
                    "Tier":              tier,
                    "eDTUs / vCores":    str(capacity),
                    "Max Storage (GiB)": round(max_gib, 2),
                    "Allocated (GiB)":   alloc,
                    "Used (GiB)":        used,
                    "Databases":         db_count,
                    "Zone Redundant":    getattr(pool, "zone_redundant", False) or False,
                })
        if verbose:
            log.info("SQL Elastic Pools %s: %d pools", sub_name, len(rows))
    except Exception as exc:
        log.warning("SQL Elastic Pools %s: %s", sub_name, exc)
    return rows


def collect_sql_vms(credential, sub_id, sub_name, verbose=False):
    """Collect VMs that have SQL Server installed via the SQL VM extension."""
    rows = []
    sql_vm_names = set()
    if _SqlVMClient is None:
        return rows, sql_vm_names
    try:
        client = _SqlVMClient(credential, sub_id)
        for sqlvm in safe_list(client.sql_virtual_machines.list()):
            rg      = rg_from_id(sqlvm.id)
            vm_name = sqlvm.virtual_machine_resource_id.split("/")[-1] if sqlvm.virtual_machine_resource_id else sqlvm.name
            sql_vm_names.add(vm_name.lower())
            rows.append({
                "Subscription":       sub_name,
                "VM Name":            vm_name,
                "Resource Group":     rg,
                "Location":           sqlvm.location or "",
                "SQL Image Offer":    getattr(sqlvm, "sql_image_offer", "") or "",
                "SQL Image SKU":      str(getattr(sqlvm, "sql_image_sku", "") or ""),
                "SQL Management":     str(getattr(sqlvm, "sql_management", "") or ""),
                "License Type":       str(getattr(sqlvm, "sql_server_license_type", "") or ""),
                "Patching Day":       str(getattr(getattr(sqlvm, "auto_patching_settings", None), "day_of_week", "") or ""),
                "Backup Enabled":     str(getattr(getattr(sqlvm, "auto_backup_settings", None), "enable", False) or False),
            })
        if verbose:
            log.info("SQL VMs %s: %d found", sub_name, len(rows))
    except Exception as exc:
        log.warning("SQL VMs %s: %s", sub_name, exc)
    return rows, sql_vm_names


def collect_storage(credential, sub_id, sub_name, verbose=False):
    """Collect Storage Accounts with blob and file sizes from Azure Monitor."""
    rows = []
    if _StorageClient is None:
        return rows
    try:
        storage = _StorageClient(credential, sub_id)
        monitor = _MonitorClient(credential, sub_id) if _MonitorClient else None

        for acct in safe_list(storage.storage_accounts.list()):
            rg          = rg_from_id(acct.id)
            sku         = acct.sku.name if acct.sku else ""
            kind        = str(acct.kind or "")
            https_only  = acct.enable_https_traffic_only or False
            public_blob = acct.allow_blob_public_access
            enc_key     = ""

            if acct.encryption:
                ksp = acct.encryption.key_source
                enc_key = str(ksp or "")

            # Determine service type from account kind
            kind_map = {
                "FileStorage":       "Azure Files (Premium)",
                "BlobStorage":       "Blob Storage",
                "BlockBlobStorage":  "Block Blob (Premium)",
                "Storage":           "General Purpose v1",
                "StorageV2":         "General Purpose v2",
            }
            svc_type = kind_map.get(kind, kind)

            # Soft delete retention (blob and file)
            blob_soft_delete = ""
            file_soft_delete = ""
            try:
                bp = storage.blob_services.get_service_properties(rg, acct.name)
                if bp.delete_retention_policy and bp.delete_retention_policy.enabled:
                    blob_soft_delete = bp.delete_retention_policy.days or ""
            except Exception:
                pass
            try:
                fp = storage.file_services.get_service_properties(rg, acct.name)
                if fp.share_delete_retention_policy and fp.share_delete_retention_policy.enabled:
                    file_soft_delete = fp.share_delete_retention_policy.days or ""
            except Exception:
                pass

            # Blob & File sizes via Azure Monitor
            blob_id   = f"{acct.id}/blobServices/default"
            file_id   = f"{acct.id}/fileServices/default"
            blob_b    = storage_metric(monitor, blob_id, "BlobCapacity", "microsoft.storage/storageaccounts/blobservices")
            file_b    = storage_metric(monitor, file_id, "FileCapacity", "microsoft.storage/storageaccounts/fileservices")
            blob_gib  = round(gib(blob_b), 4) if blob_b is not None else "N/A"
            file_gib  = round(gib(file_b), 4) if file_b is not None else "N/A"
            total_b   = (blob_b or 0) + (file_b or 0)
            total_gib = round(gib(total_b), 4) if total_b else "N/A"
            total_tib = round(tib(total_b), 6) if total_b else "N/A"

            # Blob capacity broken down by tier (single Monitor call)
            tiers     = blob_capacity_by_tier(monitor, blob_id)
            hot_gib   = round(gib(tiers.get("Hot",  0)), 4) if tiers else "N/A"
            cool_gib  = round(gib(tiers.get("Cool", 0)), 4) if tiers else "N/A"
            cold_gib  = round(gib(tiers.get("Cold", 0)), 4) if tiers else "N/A"
            arch_gib  = round(gib(tiers.get("Archive", 0)), 4) if tiers else "N/A"

            # Derive active services — use kind as primary truth, metrics to refine
            has_blob = isinstance(blob_gib, float) and blob_gib > 0
            has_file = isinstance(file_gib, float) and file_gib > 0

            if kind in ("BlobStorage", "BlockBlobStorage"):
                active_services = "Blob"
            elif kind == "FileStorage":
                active_services = "Files"
            elif kind in ("StorageV2", "Storage"):
                # Use metrics if available to be specific, otherwise say both are possible
                if has_blob and has_file:
                    active_services = "Blob + Files"
                elif has_blob:
                    active_services = "Blob"
                elif has_file:
                    active_services = "Files"
                else:
                    active_services = "Blob / Files"   # GPv2 supports both; metrics not yet available
            else:
                active_services = "Blob / Files"

            # Refine service type label for General Purpose accounts where we know usage
            if kind == "StorageV2":
                if active_services == "Files":
                    svc_type = "Azure Files (GPv2)"
                elif active_services == "Blob":
                    svc_type = "Blob Storage (GPv2)"
                elif active_services == "Blob + Files":
                    svc_type = "General Purpose v2 (Blob + Files)"

            tags_ = acct.tags or {}
            rows.append({
                "Subscription":           sub_name,
                "Name":                   acct.name,
                "Resource Group":         rg,
                "Location":               acct.location or "",
                "Service Type":           svc_type,
                "SKU":                    sku,
                "Active Services":        active_services,
                "HTTPS Only":             https_only,
                "Public Blob Access":     str(public_blob) if public_blob is not None else "Unknown",
                "Encryption Key":         enc_key,
                "Blob Soft Delete (days)":blob_soft_delete,
                "File Soft Delete (days)":file_soft_delete,
                "Blob Size (GiB)":        blob_gib,
                "Blob Hot (GiB)":         hot_gib,
                "Blob Cool (GiB)":        cool_gib,
                "Blob Cold (GiB)":        cold_gib,
                "Blob Archive (GiB)":     arch_gib,
                "File Size (GiB)":        file_gib,
                "Total Size (GiB)":       total_gib,
                "Total Size (TiB)":       total_tib,
                "Environment":            tag(tags_, "Environment", "env"),
            })

        if verbose:
            log.info("Storage %s: %d accounts", sub_name, len(rows))
    except Exception as exc:
        log.warning("Storage %s: %s", sub_name, exc)
    return rows


def collect_file_shares(credential, sub_id, sub_name, verbose=False):
    """Collect all Azure File Shares across every Storage Account in the subscription."""
    rows = []
    if _StorageClient is None:
        return rows
    try:
        storage = _StorageClient(credential, sub_id)
        for acct in safe_list(storage.storage_accounts.list()):
            rg       = rg_from_id(acct.id)
            location = acct.location or ""
            sku      = acct.sku.name if acct.sku else ""
            # SMB UNC path root for this account
            smb_root = f"\\\\{acct.name}.file.core.windows.net"

            try:
                shares = safe_list(storage.file_shares.list(rg, acct.name))
            except Exception:
                continue  # account type doesn't support file shares (e.g. ADLS Gen2, BlobStorage)

            for share in shares:
                quota_gib  = share.share_quota or 0          # provisioned size in GiB
                used_bytes = share.share_usage_bytes or 0    # actual used bytes
                used_gib   = round(gib(used_bytes), 4) if used_bytes else "N/A"
                protocols  = str(share.enabled_protocols or "SMB")
                tier       = str(share.access_tier or "")
                last_mod   = str(share.last_modified_time.date()) if share.last_modified_time else ""

                # Build UNC / mount path
                if "NFS" in protocols.upper():
                    mount_path = f"{acct.name}.file.core.windows.net:/{acct.name}/{share.name}"
                else:
                    mount_path = f"{smb_root}\\{share.name}"

                rows.append({
                    "Subscription":      sub_name,
                    "Storage Account":   acct.name,
                    "Share Name":        share.name,
                    "Resource Group":    rg,
                    "Location":          location,
                    "Protocol":          protocols,
                    "Access Tier":       tier,
                    "Quota (GiB)":       quota_gib,
                    "Used (GiB)":        used_gib,
                    "Storage SKU":       sku,
                    "Last Modified":     last_mod,
                    "Mount / UNC Path":  mount_path,
                })

        if verbose:
            log.info("File Shares %s: %d shares", sub_name, len(rows))
    except Exception as exc:
        log.warning("File Shares %s: %s", sub_name, exc)
    return rows


def collect_netapp(credential, sub_id, sub_name, verbose=False):
    """Collect Azure NetApp Files accounts, pools, and volumes with mount paths."""
    rows = []
    if _NetAppClient is None:
        return rows
    try:
        na = _NetAppClient(credential, sub_id)

        # Try subscription-level list; fall back to resource-group iteration
        try:
            accounts = safe_list(na.accounts.list_by_subscription())
        except AttributeError:
            accounts = safe_list(na.accounts.list())

        for acct in accounts:
            rg_a = rg_from_id(acct.id)
            for pool in safe_list(na.pools.list(rg_a, acct.name)):
                pool_size_gib  = (pool.size or 0) // 1_073_741_824
                svc_level      = str(pool.service_level or "")

                for vol in safe_list(na.volumes.list(rg_a, acct.name, pool.name)):
                    quota_gib   = (vol.usage_threshold or 0) // 1_073_741_824
                    used_bytes  = getattr(vol, "actual_throughput_mibps", None)  # not bytes, just check attr exists
                    # Actual used size comes from volume_usage_bytes if available
                    used_b      = getattr(vol, "volume_usage_bytes", None) or getattr(vol, "used_bytes", None)
                    used_gib    = round(gib(used_b), 4) if used_b else "N/A"
                    protocols   = ",".join(vol.protocol_types or [])
                    throughput  = getattr(vol, "throughput_mibps", None) or getattr(vol, "actual_throughput_mibps", None)
                    subnet      = (vol.subnet_id or "").split("/")[-1] if vol.subnet_id else ""
                    snap_policy = vol.snapshot_policy_id.split("/")[-1] if vol.snapshot_policy_id else "None"
                    snap_enabled= str(getattr(vol, "snapshot_directory_visible", ""))
                    vol_path    = getattr(vol, "creation_token", vol.name)  # the volume path / junction path

                    # Mount target IP(s)
                    mount_ip    = ""
                    mount_path  = ""
                    targets     = getattr(vol, "mount_targets", None) or []
                    if targets:
                        ips = [t.ip_address for t in targets if getattr(t, "ip_address", None)]
                        if ips:
                            mount_ip = ips[0]
                            if "NFSv" in protocols or "NFS" in protocols.upper():
                                mount_path = f"{mount_ip}:/{vol_path}"
                            elif "SMB" in protocols.upper() or "CIFS" in protocols.upper():
                                smb_fqdn = getattr(targets[0], "smb_server_fqdn", "") or mount_ip
                                mount_path = f"\\\\{smb_fqdn}\\{vol_path}"
                            else:
                                mount_path = f"{mount_ip}:/{vol_path}"

                    rows.append({
                        "Subscription":    sub_name,
                        "Account":         acct.name,
                        "Pool":            pool.name,
                        "Volume":          vol.name,
                        "Volume Path":     vol_path,
                        "Resource Group":  rg_a,
                        "Location":        vol.location or "",
                        "Service Level":   svc_level,
                        "Protocol":        protocols,
                        "Quota (GiB)":     quota_gib,
                        "Used (GiB)":      used_gib,
                        "Pool Size (GiB)": pool_size_gib,
                        "Throughput (MiB/s)": throughput or "",
                        "Mount Target IP": mount_ip,
                        "Mount Path":      mount_path,
                        "Subnet":          subnet,
                        "Snapshot Policy": snap_policy,
                        "Snapshot Dir":    snap_enabled,
                    })

        if verbose:
            log.info("NetApp %s: %d volumes", sub_name, len(rows))
    except Exception as exc:
        log.warning("NetApp %s: %s", sub_name, exc)
    return rows


def collect_cosmosdb(credential, sub_id, sub_name, verbose=False):
    """Collect Cosmos DB accounts with live usage metrics."""
    rows = []
    if _CosmosClient is None:
        return rows
    try:
        cosmos  = _CosmosClient(credential, sub_id)
        monitor = _MonitorClient(credential, sub_id) if _MonitorClient else None

        def _cosmos_metrics(resource_id):
            """Return (data_gib, index_gib, doc_count, partition_count) from Monitor."""
            if not monitor:
                return None, None, None, None
            try:
                end   = datetime.datetime.now(datetime.timezone.utc)
                start = end - datetime.timedelta(days=1)
                ts    = f"{start.strftime('%Y-%m-%dT%H:%M:%SZ')}/{end.strftime('%Y-%m-%dT%H:%M:%SZ')}"
                res   = monitor.metrics.list(
                    resource_id, timespan=ts, interval="PT1H",
                    metricnames="DataUsage,IndexUsage,DocumentCount,PhysicalPartitionCount",
                    aggregation="Maximum",
                    metricnamespace="microsoft.documentdb/databaseaccounts",
                )
                vals = {}
                for m in res.value:
                    key = m.name.value.lower()
                    for series in m.timeseries:
                        for dp in reversed(series.data):
                            if dp.maximum is not None:
                                vals[key] = dp.maximum
                                break
                data_gib  = round(vals.get("datausage", 0) / 1_073_741_824, 4) if "datausage" in vals else None
                index_gib = round(vals.get("indexusage", 0) / 1_073_741_824, 4) if "indexusage" in vals else None
                doc_count = int(vals["documentcount"]) if "documentcount" in vals else None
                part_cnt  = int(vals["physicalpartitioncount"]) if "physicalpartitioncount" in vals else None
                return data_gib, index_gib, doc_count, part_cnt
            except Exception:
                return None, None, None, None

        for acct in safe_list(cosmos.database_accounts.list()):
            rg          = rg_from_id(acct.id)
            kind        = str(acct.kind or "")
            consistency = str(acct.consistency_policy.default_consistency_level if acct.consistency_policy else "")
            geo_redund  = acct.enable_multiple_write_locations or False
            locations   = ",".join(l.location_name for l in (acct.locations or []))
            backup_mode = ""
            backup_ret  = ""
            if acct.backup_policy:
                backup_mode = str(getattr(acct.backup_policy, "type", "") or "")
                if hasattr(acct.backup_policy, "periodic_mode_properties") and acct.backup_policy.periodic_mode_properties:
                    backup_ret = str(acct.backup_policy.periodic_mode_properties.backup_retention_interval_in_hours or "")
            data_gib, index_gib, doc_count, part_cnt = _cosmos_metrics(acct.id)
            rows.append({
                "Subscription":       sub_name,
                "Name":               acct.name,
                "Resource Group":     rg,
                "Locations":          locations,
                "API Kind":           kind,
                "Consistency":        consistency,
                "Multi-Region Write": geo_redund,
                "Backup Mode":        backup_mode,
                "Backup Retention":   backup_ret,
                "Public Access":      str(acct.public_network_access or ""),
                "Data (GiB)":         data_gib,
                "Index (GiB)":        index_gib,
                "Document Count":     doc_count,
                "Partitions":         part_cnt,
            })
        if verbose:
            log.info("CosmosDB %s: %d accounts", sub_name, len(rows))
    except Exception as exc:
        log.warning("CosmosDB %s: %s", sub_name, exc)
    return rows


def collect_synapse(credential, sub_id, sub_name, verbose=False):
    """Collect Synapse Analytics workspaces and dedicated SQL pools."""
    rows = []
    if _SynapseClient is None:
        return rows
    try:
        syn = _SynapseClient(credential, sub_id)
        for ws in safe_list(syn.workspaces.list()):
            rg_w = rg_from_id(ws.id)
            for pool in safe_list(syn.sql_pools.list_by_workspace(rg_w, ws.name)):
                rows.append({
                    "Subscription":    sub_name,
                    "Workspace":       ws.name,
                    "SQL Pool":        pool.name,
                    "Resource Group":  rg_w,
                    "Location":        pool.location or "",
                    "SKU":             pool.sku.name if pool.sku else "",
                    "Status":          str(pool.status or ""),
                    "Collation":       pool.collation or "",
                    "Geo-Backup":      str(pool.storage_account_type or ""),
                })
        if verbose:
            log.info("Synapse %s: %d pools", sub_name, len(rows))
    except Exception as exc:
        log.warning("Synapse %s: %s", sub_name, exc)
    return rows


def collect_aks(credential, sub_id, sub_name, verbose=False):
    """Collect AKS clusters and node pools."""
    rows = []
    if _AKSClient is None:
        return rows
    try:
        aks = _AKSClient(credential, sub_id)
        for cluster in safe_list(aks.managed_clusters.list()):
            rg      = rg_from_id(cluster.id)
            version = cluster.kubernetes_version or ""
            rbac    = cluster.enable_rbac or False
            network = ""
            if cluster.network_profile:
                network = str(cluster.network_profile.network_plugin or "")

            node_pools   = cluster.agent_pool_profiles or []
            total_nodes  = sum(p.count or 0 for p in node_pools)
            max_nodes    = sum(p.max_count or p.count or 0 for p in node_pools)
            pool_names   = ",".join(p.name for p in node_pools)
            vm_sizes     = ",".join(dict.fromkeys(p.vm_size for p in node_pools if p.vm_size))
            os_disk_sizes= ",".join(
                f"{p.name}:{p.os_disk_size_gb}GiB"
                for p in node_pools if (p.os_disk_size_gb or 0) > 0
            )
            autoscale_on = any(getattr(p, "enable_auto_scaling", False) for p in node_pools)

            rows.append({
                "Subscription":      sub_name,
                "Cluster Name":      cluster.name,
                "Resource Group":    rg,
                "Location":          cluster.location or "",
                "K8s Version":       version,
                "Node Pools":        pool_names,
                "Current Nodes":     total_nodes,
                "Max Nodes":         max_nodes,
                "Autoscale":         autoscale_on,
                "Node VM Sizes":     vm_sizes,
                "OS Disk Sizes":     os_disk_sizes,
                "Network Plugin":    network,
                "RBAC Enabled":      rbac,
            })
        if verbose:
            log.info("AKS %s: %d clusters", sub_name, len(rows))
    except Exception as exc:
        log.warning("AKS %s: %s", sub_name, exc)
    return rows


def collect_container_instances(credential, sub_id, sub_name, verbose=False):
    """Collect Azure Container Instances (container groups)."""
    rows = []
    if _ACIClient is None:
        return rows
    try:
        aci = _ACIClient(credential, sub_id)
        for cg in safe_list(aci.container_groups.list()):
            rg         = rg_from_id(cg.id)
            os_type    = str(cg.os_type or "")
            containers = len(cg.containers or [])
            state      = str(cg.instance_view.state if cg.instance_view else "")
            ip_addr    = ""
            if cg.ip_address:
                ip_addr = cg.ip_address.ip or ""
            total_cpu  = sum((c.resources.requests.cpu if c.resources and c.resources.requests else 0)
                             for c in (cg.containers or []))
            total_mem  = sum((c.resources.requests.memory_in_gb if c.resources and c.resources.requests else 0)
                             for c in (cg.containers or []))
            rows.append({
                "Subscription":   sub_name,
                "Name":           cg.name,
                "Resource Group": rg,
                "Location":       cg.location or "",
                "OS Type":        os_type,
                "Containers":     containers,
                "CPU (cores)":    round(total_cpu, 2),
                "Memory (GiB)":   round(total_mem, 2),
                "State":          state,
                "IP Address":     ip_addr,
            })
        if verbose:
            log.info("ACI %s: %d container groups", sub_name, len(rows))
    except Exception as exc:
        log.warning("ACI %s: %s", sub_name, exc)
    return rows


def collect_functions(credential, sub_id, sub_name, verbose=False):
    """Collect Azure Function Apps."""
    rows = []
    if _WebClient is None:
        return rows
    try:
        web = _WebClient(credential, sub_id)
        for app in safe_list(web.web_apps.list()):
            if not app.kind or "functionapp" not in app.kind.lower():
                continue
            rg      = rg_from_id(app.id)
            runtime = ""
            if app.site_config and app.site_config.linux_fx_version:
                runtime = app.site_config.linux_fx_version
            elif app.site_config and app.site_config.windows_fx_version:
                runtime = app.site_config.windows_fx_version
            os_type = "Linux" if app.reserved else "Windows"
            plan    = app.server_farm_id.split("/")[-1] if app.server_farm_id else ""
            rows.append({
                "Subscription":   sub_name,
                "Name":           app.name,
                "Resource Group": rg,
                "Location":       app.location or "",
                "Runtime":        runtime,
                "OS Type":        os_type,
                "App Service Plan":plan,
                "State":          str(app.state or ""),
                "Kind":           app.kind or "",
            })
        if verbose:
            log.info("Functions %s: %d apps", sub_name, len(rows))
    except Exception as exc:
        log.warning("Functions %s: %s", sub_name, exc)
    return rows


def collect_avd(credential, sub_id, sub_name, verbose=False):
    """Collect Azure Virtual Desktop host pools and session hosts."""
    rows = []
    if _AVDClient is None:
        return rows
    try:
        avd = _AVDClient(credential, sub_id)
        for hp in safe_list(avd.host_pools.list()):
            rg          = rg_from_id(hp.id)
            pool_type   = str(hp.host_pool_type or "")
            lb_type     = str(hp.load_balancer_type or "")
            max_sessions= hp.max_session_limit or 0
            # Count session hosts
            sh_count = 0
            try:
                sh_count = sum(1 for _ in avd.session_hosts.list_by_host_pool(rg, hp.name))
            except Exception:
                pass
            rows.append({
                "Subscription":    sub_name,
                "Host Pool":       hp.name,
                "Resource Group":  rg,
                "Location":        hp.location or "",
                "Pool Type":       pool_type,
                "Load Balancer":   lb_type,
                "Max Sessions":    max_sessions,
                "Session Hosts":   sh_count,
            })
        if verbose:
            log.info("AVD %s: %d host pools", sub_name, len(rows))
    except Exception as exc:
        log.warning("AVD %s: %s", sub_name, exc)
    return rows


def collect_redis(credential, sub_id, sub_name, verbose=False):
    """Collect Azure Cache for Redis instances."""
    rows = []
    if _RedisClient is None:
        return rows
    try:
        redis = _RedisClient(credential, sub_id)
        for r in safe_list(redis.redis.list_by_subscription()):
            rg         = rg_from_id(r.id)
            sku        = ""
            capacity   = 0
            if r.sku:
                sku      = f"{r.sku.family}{r.sku.capacity} ({r.sku.name})"
                capacity = r.sku.capacity
            tls_only   = r.minimum_tls_version or ""
            non_ssl    = r.enable_non_ssl_port or False
            version    = r.redis_version or ""
            geo_linked = bool(getattr(r, "linked_servers", None))
            rows.append({
                "Subscription":    sub_name,
                "Name":            r.name,
                "Resource Group":  rg,
                "Location":        r.location or "",
                "SKU":             sku,
                "Capacity":        capacity,
                "Redis Version":   version,
                "Min TLS":         str(tls_only),
                "Non-SSL Port":    non_ssl,
                "Geo-Replication": geo_linked,
                "Host Name":       r.host_name or "",
            })
        if verbose:
            log.info("Redis %s: %d instances", sub_name, len(rows))
    except Exception as exc:
        log.warning("Redis %s: %s", sub_name, exc)
    return rows


def collect_backup(credential, sub_id, sub_name, verbose=False):
    """Collect Recovery Services Vaults and protected backup items."""
    vault_rows = []
    plan_rows  = []
    if _RecoveryClient is None:
        return vault_rows, plan_rows
    try:
        recovery = _RecoveryClient(credential, sub_id)
        bk_client = _BackupClient(credential, sub_id) if _BackupClient else None

        for vault in safe_list(recovery.vaults.list_by_subscription_id()):
            rg    = rg_from_id(vault.id)
            redund= ""
            if vault.properties and vault.properties.redundancy_settings:
                redund = str(vault.properties.redundancy_settings.storage_type or "")
            elif vault.sku:
                redund = vault.sku.name or ""

            item_count = 0
            if bk_client:
                try:
                    items = safe_list(bk_client.backup_protected_items.list(vault.name, rg))
                    item_count = len(items)
                    for item in items:
                        props       = item.properties
                        policy_id   = getattr(props, "policy_id", "") if props else ""
                        policy_name = policy_id.split("/")[-1] if policy_id else ""
                        plan_rows.append({
                            "Subscription":     sub_name,
                            "Vault":            vault.name,
                            "Protected Item":   item.name.split(";")[-1] if item.name else "",
                            "Resource Group":   rg,
                            "Location":         vault.location or "",
                            "Item Type":        str(getattr(props, "workload_type", "") if props else ""),
                            "Protection Status":str(getattr(props, "protection_status", "") if props else ""),
                            "Last Backup":      str(getattr(props, "last_backup_time", "") if props else ""),
                            "Policy Name":      policy_name,
                        })
                except Exception:
                    pass

            vault_rows.append({
                "Subscription":    sub_name,
                "Vault Name":      vault.name,
                "Resource Group":  rg,
                "Location":        vault.location or "",
                "Redundancy":      redund,
                "Protected Items": item_count,
            })

        if verbose:
            log.info("Backup %s: %d vaults, %d items", sub_name, len(vault_rows), len(plan_rows))
    except Exception as exc:
        log.warning("Backup %s: %s", sub_name, exc)
    return vault_rows, plan_rows


def collect_backup_sql_items(credential, sub_id, sub_name, verbose=False):
    """Collect SQL workload protected items (SQLDataBase + SQLInstance) from all vaults."""
    rows = []
    if _RecoveryClient is None or _BackupClient is None:
        return rows
    try:
        recovery  = _RecoveryClient(credential, sub_id)
        bk_client = _BackupClient(credential, sub_id)
        for vault in safe_list(recovery.vaults.list_by_subscription_id()):
            rg = rg_from_id(vault.id)
            try:
                items = safe_list(
                    bk_client.backup_protected_items.list(
                        vault.name, rg,
                        filter="backupManagementType eq 'AzureWorkload'"
                    )
                )
            except Exception:
                continue
            for item in items:
                props        = item.properties
                workload_type= str(getattr(props, "workload_type", "") if props else "")
                if workload_type.lower() not in ("sqldatabase", "sqlinstance", "sqldataguard",
                                                  "sqlavailabilitygrouplisten", ""):
                    if "sql" not in workload_type.lower():
                        continue
                policy_id   = getattr(props, "policy_id", "") if props else ""
                policy_name = policy_id.split("/")[-1] if policy_id else ""
                # item name format: "SQLDataBase;MSSQLSERVER;dbname" or similar
                parts = (item.name or "").split(";")
                db_name  = parts[-1] if len(parts) >= 2 else item.name or ""
                srv_name = parts[1]  if len(parts) >= 3 else ""
                rows.append({
                    "Subscription":      sub_name,
                    "Vault":             vault.name,
                    "Server / Instance": srv_name,
                    "Database":          db_name,
                    "Resource Group":    rg,
                    "Location":          vault.location or "",
                    "Workload Type":     workload_type,
                    "Protection Status": str(getattr(props, "protection_status", "") if props else ""),
                    "Last Backup":       str(getattr(props, "last_backup_time", "") if props else ""),
                    "Last Backup Status":str(getattr(props, "last_backup_status", "") if props else ""),
                    "Policy Name":       policy_name,
                })
        if verbose:
            log.info("Backup SQL items %s: %d items", sub_name, len(rows))
    except Exception as exc:
        log.warning("Backup SQL items %s: %s", sub_name, exc)
    return rows


def collect_backup_policies(credential, sub_id, sub_name, verbose=False):
    """Collect backup policies (VM and SQL) from all Recovery Services Vaults."""
    rows = []
    if _RecoveryClient is None or _BackupClient is None:
        return rows
    try:
        recovery  = _RecoveryClient(credential, sub_id)
        bk_client = _BackupClient(credential, sub_id)
        for vault in safe_list(recovery.vaults.list_by_subscription_id()):
            rg = rg_from_id(vault.id)
            try:
                policies = safe_list(bk_client.backup_policies.list(vault.name, rg))
            except Exception:
                continue
            for pol in policies:
                props   = pol.properties
                bm_type = str(getattr(props, "backup_management_type", "") if props else "")
                # Schedule details
                schedule_freq = ""
                retention_days= ""
                weekly_ret    = ""
                monthly_ret   = ""
                yearly_ret    = ""
                try:
                    sp = getattr(props, "schedule_policy", None)
                    if sp:
                        schedule_freq = str(getattr(sp, "schedule_run_frequency", "") or "")
                except Exception:
                    pass
                try:
                    rp = getattr(props, "retention_policy", None)
                    if rp:
                        daily = getattr(rp, "daily_schedule", None)
                        if daily:
                            dr = getattr(daily, "retention_duration", None)
                            if dr:
                                retention_days = str(getattr(dr, "count", "") or "")
                        weekly_ret  = str(getattr(getattr(rp, "weekly_schedule",  None), "retention_times", [None])[0] or "") if getattr(rp, "weekly_schedule",  None) else ""
                        monthly_ret = "Yes" if getattr(rp, "monthly_schedule", None) else ""
                        yearly_ret  = "Yes" if getattr(rp, "yearly_schedule",  None) else ""
                except Exception:
                    pass
                # For AzureWorkload (SQL) policies, retention lives in sub-protection policies
                if not retention_days:
                    try:
                        for sp in (getattr(props, "sub_protection_policy", None) or []):
                            rp = getattr(sp, "retention_policy", None)
                            if rp:
                                daily = getattr(rp, "daily_schedule", None)
                                if daily:
                                    dr = getattr(daily, "retention_duration", None)
                                    if dr:
                                        retention_days = str(getattr(dr, "count", "") or "")
                                        break
                    except Exception:
                        pass
                rows.append({
                    "Subscription":    sub_name,
                    "Vault":           vault.name,
                    "Policy Name":     pol.name or "",
                    "Resource Group":  rg,
                    "Location":        vault.location or "",
                    "Type":            bm_type,
                    "Schedule":        schedule_freq,
                    "Daily Retention": retention_days,
                    "Weekly Retention":weekly_ret,
                    "Monthly Enabled": monthly_ret,
                    "Yearly Enabled":  yearly_ret,
                })
        if verbose:
            log.info("Backup policies %s: %d policies", sub_name, len(rows))
    except Exception as exc:
        log.warning("Backup policies %s: %s", sub_name, exc)
    return rows


def collect_cloud_spend(credential, sub_id, sub_name, verbose=False):
    """Collect monthly Azure spend by service category — current month and previous month."""
    rows = []
    if _CostClient is None:
        return rows
    try:
        import importlib
        models = importlib.import_module("azure.mgmt.costmanagement.models")
        client = _CostClient(credential)
        scope  = f"/subscriptions/{sub_id}"

        today      = datetime.datetime.now(datetime.timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        # Current month: 1st of this month → today
        cur_start  = today.replace(day=1)
        cur_end    = today
        # Previous month: 1st → last day of last month
        prev_end   = cur_start - datetime.timedelta(days=1)
        prev_start = prev_end.replace(day=1)

        def _query_spend(start, end, include_family=True):
            grouping = [models.QueryGrouping(type="Dimension", name="ServiceName")]
            if include_family:
                grouping.append(models.QueryGrouping(type="Dimension", name="ServiceFamily"))
            params = models.QueryDefinition(
                type="ActualCost",
                timeframe="Custom",
                time_period=models.QueryTimePeriod(from_property=start, to=end),
                dataset=models.QueryDataset(
                    granularity="None",
                    aggregation={"totalCost": models.QueryAggregation(name="Cost", function="Sum")},
                    grouping=grouping,
                ),
            )
            result = client.query.usage(scope, params)
            col_idx  = {c.name.lower(): i for i, c in enumerate(result.columns or [])}
            cost_i   = col_idx.get("cost", col_idx.get("pretaxcost", 0))
            svc_i    = col_idx.get("servicename", 1)
            fam_i    = col_idx.get("servicefamily", 2) if include_family else None
            cur_i    = col_idx.get("currency", None)
            out = {}
            for row in (result.rows or []):
                svc      = str(row[svc_i])  if svc_i  < len(row) else "Other"
                fam      = (str(row[fam_i]) if fam_i is not None and fam_i < len(row) else "")
                cost     = float(row[cost_i]) if cost_i < len(row) else 0.0
                currency = str(row[cur_i])  if cur_i is not None and cur_i < len(row) else "USD"
                out[svc] = (round(cost, 4), fam, currency)
            return out

        # Some subscription types (e.g. Visual Studio/MSDN) don't support ServiceFamily
        try:
            cur_spend  = _query_spend(cur_start, cur_end,   include_family=True)
            prev_spend = _query_spend(prev_start, prev_end, include_family=True)
        except Exception:
            cur_spend  = _query_spend(cur_start, cur_end,   include_family=False)
            prev_spend = _query_spend(prev_start, prev_end, include_family=False)

        all_services = sorted(set(cur_spend) | set(prev_spend))
        cur_label  = cur_start.strftime("%b %Y")
        prev_label = prev_start.strftime("%b %Y")

        for svc in all_services:
            cur_cost,  fam,  currency = cur_spend.get(svc,  (0.0, "", "USD"))
            prev_cost, _,    _        = prev_spend.get(svc, (0.0, "", "USD"))
            if prev_cost > 0:
                delta_pct = round(((cur_cost - prev_cost) / prev_cost) * 100, 1)
            else:
                delta_pct = ""
            rows.append({
                "Subscription":         sub_name,
                "Service":              svc,
                "Service Family":       fam,
                f"Cost ({cur_label})":  cur_cost,
                f"Cost ({prev_label})": prev_cost,
                "MoM Change (%)":       delta_pct,
                "Currency":             currency,
            })

        if verbose:
            log.info("Cloud spend %s: %d service categories", sub_name, len(rows))
    except Exception as exc:
        log.warning("Cloud spend %s: %s", sub_name, exc)
    return rows


def collect_backup_costs(credential, sub_id, sub_name, verbose=False):
    """Collect Azure Cost Management data for Recovery Services Vaults (last 30 days)."""
    rows = []
    if _CostClient is None:
        return rows
    try:
        import importlib
        models = importlib.import_module("azure.mgmt.costmanagement.models")
        client = _CostClient(credential)
        scope  = f"/subscriptions/{sub_id}"

        end_dt   = datetime.datetime.now(datetime.timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        start_dt = end_dt - datetime.timedelta(days=30)

        params = models.QueryDefinition(
            type="ActualCost",
            timeframe="Custom",
            time_period=models.QueryTimePeriod(
                from_property=start_dt,
                to=end_dt,
            ),
            dataset=models.QueryDataset(
                granularity="None",
                aggregation={"totalCost": models.QueryAggregation(name="Cost", function="Sum")},
                grouping=[
                    models.QueryGrouping(type="Dimension", name="ResourceId"),
                    models.QueryGrouping(type="Dimension", name="ResourceGroupName"),
                    models.QueryGrouping(type="Dimension", name="ResourceLocation"),
                ],
                filter=models.QueryFilter(
                    dimensions=models.QueryComparisonExpression(
                        name="ResourceType",
                        operator="In",
                        values=["microsoft.recoveryservices/vaults"],
                    )
                ),
            ),
        )

        result = client.query.usage(scope, params)

        # Map column names to indices
        col_idx = {c.name.lower(): i for i, c in enumerate(result.columns or [])}
        cost_i  = col_idx.get("cost", col_idx.get("pretaxcost", 0))
        rid_i   = col_idx.get("resourceid", 1)
        rg_i    = col_idx.get("resourcegroupname", 2)
        loc_i   = col_idx.get("resourcelocation", 3)
        cur_i   = col_idx.get("currency", None)

        period = f"{start_dt.strftime('%Y-%m-%d')} – {end_dt.strftime('%Y-%m-%d')}"
        for row in (result.rows or []):
            rid       = str(row[rid_i]) if rid_i < len(row) else ""
            vault_name= rid.split("/")[-1] if rid else ""
            rg        = str(row[rg_i])  if rg_i  < len(row) else ""
            loc       = str(row[loc_i]) if loc_i < len(row) else ""
            cost      = row[cost_i] if cost_i < len(row) else 0
            currency  = str(row[cur_i]) if cur_i is not None and cur_i < len(row) else "USD"
            rows.append({
                "Subscription":   sub_name,
                "Vault Name":     vault_name,
                "Resource Group": rg,
                "Location":       loc,
                "Cost (30 days)": round(float(cost), 4),
                "Currency":       currency,
                "Period":         period,
            })

        if verbose:
            log.info("Backup costs %s: %d vaults", sub_name, len(rows))
    except Exception as exc:
        log.warning("Backup costs %s: %s", sub_name, exc)
    return rows


# ─── ANONYMIZER ───────────────────────────────────────────────────────────────

import csv as _csv

# Fields to anonymize and the category prefix to use
_ANON_FIELDS = {
    "Subscription":      ("sub",   "SUB"),
    "Resource Group":    ("rg",    "RG"),
    "Name":              ("res",   "RES"),
    "VM Name":           ("vm",    "VM"),
    "Cluster Name":      ("vm",    "VM"),
    "Server / Instance": ("sql",   "SQL"),
    "Server":            ("sql",   "SQL"),
    "Managed Instance":  ("sql",   "SQL"),
    "Database Name":     ("sqldb", "SQLDB"),
    "Pool Name":         ("pool",  "POOL"),
    "Workspace":         ("syn",   "SYN"),
    "Host Pool":         ("avd",   "AVD"),
    "Vault Name":        ("vault", "VAULT"),
    "Vault":             ("vault", "VAULT"),
    "Storage Account":   ("sa",    "SA"),
    "Share Name":        ("res",   "RES"),
    "Protected Item":    ("res",   "RES"),
    "SQL Pool":          ("pool",  "POOL"),
}


class Anonymizer:
    """Replace real Azure resource names with opaque codes; export a reversible mapping CSV."""

    def __init__(self):
        self._maps     = defaultdict(dict)
        self._counters = defaultdict(int)

    def map(self, category, prefix, value):
        if not value or not str(value).strip():
            return value
        key = str(value)
        if key not in self._maps[category]:
            self._counters[category] += 1
            self._maps[category][key] = f"{prefix}-{self._counters[category]:04d}"
        return self._maps[category][key]

    def apply(self, rows):
        """Return a copy of rows with identifying fields replaced by codes."""
        out = []
        for row in rows:
            r = dict(row)
            for field, (cat, prefix) in _ANON_FIELDS.items():
                if field in r:
                    r[field] = self.map(cat, prefix, r[field])
            out.append(r)
        return out

    def save(self, path):
        """Write the reversible mapping to a CSV file."""
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = _csv.writer(f)
            w.writerow(["Category", "Code", "Original Value"])
            for cat, mapping in sorted(self._maps.items()):
                for real, code in sorted(mapping.items(), key=lambda x: x[1]):
                    w.writerow([cat, code, real])
        log.info("Anonymization mapping saved: %s", path)


# ─── SHEET BUILDERS ───────────────────────────────────────────────────────────

def _new_sheet(wb, title):
    ws = wb.create_sheet(title=title[:31])
    ws.sheet_view.showGridLines = False
    return ws

def _header_row(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill   = HEADER_FILL
        c.font   = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[row].height = 22
    return row + 1

def _data_row(ws, values, row, alt=False):
    fill = ALT_FILL if alt else None
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.alignment = LEFT
        c.border    = BORDER
        c.font      = NORMAL
        if fill:
            c.fill = fill
    return row + 1

def _risk_cell(cell, value, bad_values=None, warn_values=None, good_values=None):
    """Apply RED/YELLOW/GREEN fill based on value."""
    val = str(value).lower().strip()
    if bad_values  and any(b in val for b in bad_values):  cell.fill = RED_FILL
    elif warn_values and any(w in val for w in warn_values): cell.fill = YELLOW_FILL
    elif good_values and any(g in val for g in good_values): cell.fill = GREEN_FILL

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def build_sheet_vms(wb, rows):
    ws = _new_sheet(wb, "Virtual Machines")
    headers = ["Subscription","Name","Resource Group","Location","VM Size",
               "OS Type","Power State","OS Disk (GiB)","Data Disks",
               "Data Disk (GiB)","Total Storage (GiB)","Zones",
               "Environment","Owner","Backup Tag",
               "MSSQL-INSTALLED","Backup Policy","Backup Protected"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h, "") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Power state colouring (col 7)
        ps = str(row.get("Power State","")).lower()
        ws.cell(row=r, column=7).fill = (RED_FILL if "deallocated" in ps or "stopped" in ps
                                          else GREEN_FILL if "running" in ps else YELLOW_FILL)
        # MSSQL-INSTALLED (col 16)
        ws.cell(row=r, column=16).fill = (YELLOW_FILL if row.get("MSSQL-INSTALLED") == "Yes" else PatternFill())
        # Backup Protected (col 18)
        bp = str(row.get("Backup Protected","")).lower()
        ws.cell(row=r, column=18).fill = (GREEN_FILL if bp == "yes"
                                           else RED_FILL if bp == "no"
                                           else PatternFill())
        r += 1
    _set_col_widths(ws, [18,22,20,14,18,10,14,12,10,12,14,8,14,14,14,16,18,16])
    _freeze(ws)
    return ws


def build_sheet_disks(wb, rows):
    ws = _new_sheet(wb, "Managed Disks")
    headers = ["Subscription","Name","Resource Group","Location","SKU",
               "Size (GiB)","IOPS","Throughput (MB/s)","Encryption Type",
               "Disk State","Attached To","Snapshot Coverage",
               "AKS PV","PVC Name","PVC Namespace","Environment"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Disk state (col 10)
        state = str(row.get("Disk State","")).lower()
        ws.cell(row=r, column=10).fill = (YELLOW_FILL if "unattached" in state
                                           else GREEN_FILL if "attached" in state else PatternFill())
        # Snapshot coverage (col 12)
        cov = str(row.get("Snapshot Coverage",""))
        ws.cell(row=r, column=12).fill = (RED_FILL    if "No Snapshot" in cov
                                           else RED_FILL    if "Stale"      in cov
                                           else YELLOW_FILL if "Aging"      in cov
                                           else GREEN_FILL  if cov else PatternFill())
        # AKS PV (col 13)
        if row.get("AKS PV") == "Yes":
            ws.cell(row=r, column=13).fill = PatternFill("solid", fgColor="DEEBF7")
        r += 1

    # Totals row
    if rows:
        ws.cell(row=r, column=1, value="TOTAL").font = BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        total_gib = sum(x.get("Size (GiB)",0) for x in rows if isinstance(x.get("Size (GiB)"),(int,float)))
        c = ws.cell(row=r, column=6, value=round(total_gib, 2))
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = HEADER_FILL; c.border = BORDER; c.alignment = CENTER
        ws.row_dimensions[r].height = 18

    _set_col_widths(ws, [18,24,20,14,18,10,8,14,22,14,22,18,8,22,18,14])
    _freeze(ws)
    return ws


def build_sheet_snapshots(wb, rows):
    ws = _new_sheet(wb, "Disk Snapshots")
    headers = ["Subscription","Name","Resource Group","Location","Source Disk",
               "Size (GiB)","Encryption","Created","Age (days)"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Age colouring (col 9)
        age = row.get("Age (days)","")
        if isinstance(age, int):
            ws.cell(row=r, column=9).fill = (RED_FILL if age > 365 else
                                              YELLOW_FILL if age > 90 else GREEN_FILL)
        r += 1
    _set_col_widths(ws, [18,26,20,14,26,10,22,12,10])
    _freeze(ws)
    return ws


def build_sheet_sql(wb, rows):
    ws = _new_sheet(wb, "Azure SQL")
    headers = ["Subscription","Type","Server / Instance","Database Name","Resource Group",
               "Location","SKU","Tier","Capacity","Max Storage (GiB)",
               "Allocated (GiB)","Used (GiB)","Elastic Pool",
               "PITR (Days)","LTR Weekly","LTR Monthly","LTR Yearly",
               "Backup Redundancy","Public Access","Encryption (TDE)","Availability Zone"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Public access colouring (col 19)
        pa = str(row.get("Public Access","")).lower()
        ws.cell(row=r, column=19).fill = (RED_FILL if "enabled" in pa or "true" in pa
                                           else GREEN_FILL if "disabled" in pa or "false" in pa
                                           else PatternFill())
        # PITR days (col 14) — only applicable for Azure SQL Database rows
        if str(row.get("Type","")).strip() == "Azure SQL Database":
            pitr = row.get("PITR (Days)","")
            ws.cell(row=r, column=14).fill = (RED_FILL if pitr == "" else GREEN_FILL)
        r += 1
    _set_col_widths(ws, [18,20,24,20,20,14,14,12,10,14,13,12,18,12,12,12,12,16,14,14,14])
    _freeze(ws)
    return ws


def build_sheet_sql_mi_databases(wb, rows):
    ws = _new_sheet(wb, "SQL MI Databases")
    headers = ["Subscription","Managed Instance","Database Name","Resource Group",
               "Location","License Type","Instance vCores","Instance Storage (GiB)",
               "DB Status","Collation","Created","Earliest Restore",
               "PITR (Days)","LTR Weekly","LTR Monthly","LTR Yearly"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Status colouring (col 9 — DB Status)
        status = str(row.get("DB Status","")).lower()
        ws.cell(row=r, column=9).fill = (GREEN_FILL if "online" in status
                                          else RED_FILL if "offline" in status or "error" in status
                                          else YELLOW_FILL if status else PatternFill())
        # Earliest restore (col 12) — red if empty
        er = str(row.get("Earliest Restore",""))
        ws.cell(row=r, column=12).fill = (RED_FILL if not er else PatternFill())
        r += 1
    _set_col_widths(ws, [18,26,24,20,14,14,14,18,14,20,14,16,12,14,14,12])
    _freeze(ws)
    return ws


def build_sheet_storage(wb, rows):
    ws = _new_sheet(wb, "Storage Accounts")
    headers = ["Subscription","Name","Resource Group","Location",
               "Service Type","SKU","Active Services",
               "HTTPS Only","Public Blob Access","Encryption Key",
               "Blob Soft Delete (days)","File Soft Delete (days)",
               "Blob Size (GiB)","Blob Hot (GiB)","Blob Cool (GiB)","Blob Cold (GiB)","Blob Archive (GiB)",
               "File Size (GiB)","Total Size (GiB)","Total Size (TiB)","Environment"]
    r = _header_row(ws, headers)

    # Service Type fill colours
    svc_fills = {
        "Azure Files":         PatternFill("solid", fgColor="BDD7EE"),  # blue
        "Blob Storage":        PatternFill("solid", fgColor="E2EFDA"),  # green
        "Block Blob":          PatternFill("solid", fgColor="D9E1F2"),  # indigo
        "General Purpose":     PatternFill("solid", fgColor="EDEDED"),  # grey
    }

    def svc_fill(svc_type):
        for key, fill in svc_fills.items():
            if key.lower() in svc_type.lower():
                return fill
        return PatternFill()

    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Service Type badge (col 5)
        ws.cell(row=r, column=5).fill = svc_fill(str(row.get("Service Type","")))
        ws.cell(row=r, column=5).font = Font(bold=True, name="Calibri", size=10)
        # HTTPS Only (col 8)
        ws.cell(row=r, column=8).fill = (GREEN_FILL if row.get("HTTPS Only") else RED_FILL)
        # Public Blob (col 9)
        pb = str(row.get("Public Blob Access","")).lower()
        ws.cell(row=r, column=9).fill = (RED_FILL if "true" in pb
                                          else GREEN_FILL if "false" in pb else PatternFill())
        # Soft delete — red if empty (no protection)
        ws.cell(row=r, column=11).fill = (RED_FILL if row.get("Blob Soft Delete (days)") == "" else GREEN_FILL)
        ws.cell(row=r, column=12).fill = (RED_FILL if row.get("File Soft Delete (days)") == "" else GREEN_FILL)
        # Archive tier (col 17) — yellow if data exists in archive
        arch = row.get("Blob Archive (GiB)", 0)
        if isinstance(arch, float) and arch > 0:
            ws.cell(row=r, column=17).fill = YELLOW_FILL
        r += 1

    # Totals row
    if rows:
        ws.cell(row=r, column=1, value="TOTAL").font = BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        for col, key in [(13,"Blob Size (GiB)"), (18,"File Size (GiB)"),
                         (19,"Total Size (GiB)"), (20,"Total Size (TiB)")]:
            total = sum(x.get(key,0) for x in rows if isinstance(x.get(key),(int,float)))
            c = ws.cell(row=r, column=col, value=round(total,4))
            c.font = BOLD; c.fill = HEADER_FILL
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            c.border = BORDER; c.alignment = CENTER
        ws.row_dimensions[r].height = 18

    _set_col_widths(ws, [18,24,20,14,26,18,18,10,16,24,20,20,12,12,12,12,14,12,14,14,14])
    _freeze(ws)
    return ws


def build_sheet_file_shares(wb, rows):
    ws = _new_sheet(wb, "File Shares")
    headers = ["Subscription","Storage Account","Share Name","Resource Group","Location",
               "Protocol","Access Tier","Quota (GiB)","Used (GiB)","Storage SKU",
               "Last Modified","Mount / UNC Path"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h, "") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Protocol colouring (col 6) — highlight NFS in yellow (needs different auth)
        proto = str(row.get("Protocol", "")).upper()
        ws.cell(row=r, column=6).fill = (YELLOW_FILL if "NFS" in proto else GREEN_FILL)
        # Used vs Quota — red if used GiB is N/A (no metrics yet)
        used = row.get("Used (GiB)", "")
        ws.cell(row=r, column=9).fill = (YELLOW_FILL if used == "N/A" else PatternFill())
        r += 1
    _set_col_widths(ws, [18, 26, 22, 20, 14, 10, 14, 10, 10, 18, 14, 42])
    _freeze(ws)
    return ws


def build_sheet_netapp(wb, rows):
    ws = _new_sheet(wb, "Azure NetApp Files")
    headers = ["Subscription","Account","Pool","Volume","Volume Path",
               "Resource Group","Location","Service Level","Protocol",
               "Quota (GiB)","Used (GiB)","Pool Size (GiB)","Throughput (MiB/s)",
               "Mount Target IP","Mount Path","Subnet","Snapshot Policy","Snapshot Dir"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h, "") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Protocol colouring (col 9)
        proto = str(row.get("Protocol", "")).upper()
        ws.cell(row=r, column=9).fill = (YELLOW_FILL if "SMB" in proto or "CIFS" in proto
                                          else GREEN_FILL if "NFS" in proto else PatternFill())
        # Used GiB (col 11) — yellow if N/A
        used = row.get("Used (GiB)", "")
        ws.cell(row=r, column=11).fill = (YELLOW_FILL if used == "N/A" else PatternFill())
        # Mount path (col 15) — red if empty (no mount target yet)
        mp = str(row.get("Mount Path", ""))
        ws.cell(row=r, column=15).fill = (RED_FILL if not mp else PatternFill())
        r += 1
    _set_col_widths(ws, [18, 20, 16, 18, 18, 18, 12, 14, 12, 10, 10, 12, 14, 16, 38, 16, 18, 12])
    _freeze(ws)
    return ws


def build_sheet_cosmosdb(wb, rows):
    ws = _new_sheet(wb, "Cosmos DB")
    headers = ["Subscription","Name","Resource Group","Locations","API Kind",
               "Consistency","Multi-Region Write","Backup Mode","Backup Retention","Public Access",
               "Data (GiB)","Index (GiB)","Document Count","Partitions"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        r += 1
    _set_col_widths(ws, [18,22,18,22,14,16,16,14,14,14,12,12,14,12])
    _freeze(ws)
    return ws


def build_sheet_synapse(wb, rows):
    ws = _new_sheet(wb, "Synapse Analytics")
    headers = ["Subscription","Workspace","SQL Pool","Resource Group","Location",
               "SKU","Status","Collation","Geo-Backup"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        r += 1
    _set_col_widths(ws, [18,22,20,18,14,10,12,20,14])
    _freeze(ws)
    return ws


def build_sheet_aks(wb, rows):
    ws = _new_sheet(wb, "AKS")
    headers = ["Subscription","Cluster Name","Resource Group","Location","K8s Version",
               "Node Pools","Current Nodes","Max Nodes","Autoscale",
               "Node VM Sizes","OS Disk Sizes","Network Plugin","RBAC Enabled"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Autoscale (col 9)
        ws.cell(row=r, column=9).fill = (GREEN_FILL if row.get("Autoscale") else PatternFill())
        r += 1
    _set_col_widths(ws, [18,22,20,14,12,20,12,10,10,24,26,14,12])
    _freeze(ws)
    return ws


def build_sheet_aci(wb, rows):
    ws = _new_sheet(wb, "Container Instances")
    headers = ["Subscription","Name","Resource Group","Location","OS Type",
               "Containers","CPU (cores)","Memory (GiB)","State","IP Address"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        r += 1
    _set_col_widths(ws, [18,22,20,14,10,10,10,12,12,16])
    _freeze(ws)
    return ws


def build_sheet_functions(wb, rows):
    ws = _new_sheet(wb, "Function Apps")
    headers = ["Subscription","Name","Resource Group","Location","Runtime",
               "OS Type","App Service Plan","State","Kind"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        r += 1
    _set_col_widths(ws, [18,24,20,14,20,10,20,12,14])
    _freeze(ws)
    return ws


def build_sheet_avd(wb, rows):
    ws = _new_sheet(wb, "Azure Virtual Desktop")
    headers = ["Subscription","Host Pool","Resource Group","Location",
               "Pool Type","Load Balancer","Max Sessions","Session Hosts"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        r += 1
    _set_col_widths(ws, [18,22,20,14,14,14,12,12])
    _freeze(ws)
    return ws


def build_sheet_redis(wb, rows):
    ws = _new_sheet(wb, "Redis Cache")
    headers = ["Subscription","Name","Resource Group","Location","SKU","Capacity",
               "Redis Version","Min TLS","Non-SSL Port","Geo-Replication","Host Name"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Non-SSL (col 9) — red if enabled
        ws.cell(row=r, column=9).fill = (RED_FILL if row.get("Non-SSL Port") else GREEN_FILL)
        r += 1
    _set_col_widths(ws, [18,22,20,14,18,10,12,10,12,14,28])
    _freeze(ws)
    return ws


def build_sheet_backup_vaults(wb, rows):
    ws = _new_sheet(wb, "Backup Vaults")
    headers = ["Subscription","Vault Name","Resource Group","Location","Redundancy","Protected Items"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Protected items col 6 — red if 0
        cnt = row.get("Protected Items", 0)
        ws.cell(row=r, column=6).fill = (RED_FILL if cnt == 0 else GREEN_FILL)
        r += 1
    _set_col_widths(ws, [18,24,20,14,16,14])
    _freeze(ws)
    return ws


def build_sheet_backup_items(wb, rows):
    ws = _new_sheet(wb, "Backup Protected Items")
    headers = ["Subscription","Vault","Protected Item","Resource Group","Location",
               "Item Type","Protection Status","Last Backup","Policy Name"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Protection status (col 7) colour
        ps = str(row.get("Protection Status","")).lower()
        ws.cell(row=r, column=7).fill = (GREEN_FILL if "protected" in ps and "not" not in ps
                                          else RED_FILL if "not" in ps or "error" in ps
                                          else PatternFill())
        r += 1
    _set_col_widths(ws, [18,22,28,20,14,18,16,22,22])
    _freeze(ws)
    return ws


def build_sheet_backup_sql_items(wb, rows):
    ws = _new_sheet(wb, "Backup SQL Items")
    headers = ["Subscription","Vault","Server / Instance","Database",
               "Resource Group","Location","Workload Type",
               "Protection Status","Last Backup","Last Backup Status","Policy Name"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Protection status (col 8) colour
        ps = str(row.get("Protection Status","")).lower()
        ws.cell(row=r, column=8).fill = (GREEN_FILL if "protected" in ps and "not" not in ps
                                          else RED_FILL if "not" in ps or "error" in ps
                                          else PatternFill())
        # Last backup status (col 10)
        lbs = str(row.get("Last Backup Status","")).lower()
        ws.cell(row=r, column=10).fill = (GREEN_FILL if lbs == "completed"
                                           else RED_FILL if "fail" in lbs or "error" in lbs
                                           else PatternFill())
        r += 1
    _set_col_widths(ws, [18,22,26,26,20,14,18,16,22,18,22])
    _freeze(ws)
    return ws


def build_sheet_backup_policies(wb, rows):
    ws = _new_sheet(wb, "Backup Policies")
    headers = ["Subscription","Vault","Policy Name","Resource Group","Location",
               "Type","Schedule","Daily Retention","Weekly Retention",
               "Monthly Enabled","Yearly Enabled"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        r += 1
    _set_col_widths(ws, [18,22,26,20,14,20,14,14,14,14,12])
    _freeze(ws)
    return ws


def build_sheet_backup_costs(wb, rows):
    ws = _new_sheet(wb, "Backup Costs")
    headers = ["Subscription","Vault Name","Resource Group","Location",
               "Cost (30 days)","Currency","Period"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        r += 1
    # Totals row
    if rows:
        ws.cell(row=r, column=1, value="TOTAL").font = BOLD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        total = sum(x.get("Cost (30 days)", 0) for x in rows if isinstance(x.get("Cost (30 days)"), (int, float)))
        currency = rows[0].get("Currency", "USD") if rows else "USD"
        c = ws.cell(row=r, column=5, value=round(total, 2))
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = HEADER_FILL; c.border = BORDER; c.alignment = CENTER
        ws.cell(row=r, column=6, value=currency).font = BOLD
    _set_col_widths(ws, [18,26,20,14,16,10,28])
    _freeze(ws)
    return ws


def build_sheet_cloud_spend(wb, rows):
    ws = _new_sheet(wb, "Monthly Cloud Spend")
    if not rows:
        ws.cell(row=1, column=1, value="No cost data available — ensure the account has Cost Management Reader access.")
        return ws

    # Column headers are dynamic (month names vary), so derive from first row
    fixed = ["Subscription", "Service", "Service Family"]
    cost_cols = [k for k in rows[0] if k.startswith("Cost (")]
    extra = ["MoM Change (%)", "Currency"]
    headers = fixed + cost_cols + extra

    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h, "") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # MoM change colouring — red = cost went up, green = down
        mom_col = len(fixed) + len(cost_cols) + 1
        mom = row.get("MoM Change (%)", "")
        if isinstance(mom, (int, float)):
            ws.cell(row=r, column=mom_col).fill = (RED_FILL if mom > 0 else GREEN_FILL if mom < 0 else PatternFill())
        r += 1

    # Totals row
    ws.cell(row=r, column=1, value="TOTAL").font = BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for ci, key in enumerate(cost_cols, len(fixed) + 1):
        total = sum(x.get(key, 0) for x in rows if isinstance(x.get(key), (int, float)))
        c = ws.cell(row=r, column=ci, value=round(total, 2))
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = HEADER_FILL; c.border = BORDER; c.alignment = CENTER

    widths = [18, 34, 20] + [16] * len(cost_cols) + [14, 10]
    _set_col_widths(ws, widths)
    _freeze(ws)
    return ws


def build_sheet_elastic_pools(wb, rows):
    ws = _new_sheet(wb, "SQL Elastic Pools")
    headers = ["Subscription","Server","Pool Name","Resource Group","Location",
               "SKU","Tier","eDTUs / vCores","Max Storage (GiB)",
               "Allocated (GiB)","Used (GiB)","Databases","Zone Redundant"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        r += 1
    # Totals row
    if rows:
        r2 = len(rows) + 2
        ws.cell(row=r2, column=1, value="TOTAL").font = BOLD
        ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=5)
        for col, key in [(9,"Max Storage (GiB)"), (10,"Allocated (GiB)"), (11,"Used (GiB)")]:
            total = sum(x.get(key,0) for x in rows if isinstance(x.get(key),(int,float)))
            c = ws.cell(row=r2, column=col, value=round(total,4))
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            c.fill = HEADER_FILL; c.border = BORDER; c.alignment = CENTER
    _set_col_widths(ws, [18,24,22,20,14,14,12,14,16,14,12,10,14])
    _freeze(ws)
    return ws


def build_sheet_sql_vms(wb, rows):
    ws = _new_sheet(wb, "SQL Server VMs")
    headers = ["Subscription","VM Name","Resource Group","Location",
               "SQL Image Offer","SQL Image SKU","SQL Management",
               "License Type","Patching Day","Backup Enabled"]
    r = _header_row(ws, headers)
    for i, row in enumerate(rows):
        vals = [row.get(h,"") for h in headers]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.alignment = LEFT; c.border = BORDER; c.font = NORMAL
            if i % 2 == 1: c.fill = ALT_FILL
        # Backup enabled (col 10) — red if false
        bk = str(row.get("Backup Enabled","")).lower()
        ws.cell(row=r, column=10).fill = (GREEN_FILL if "true" in bk
                                           else YELLOW_FILL if not bk
                                           else RED_FILL)
        r += 1
    _set_col_widths(ws, [18,24,20,14,22,20,18,16,14,14])
    _freeze(ws)
    return ws


# ─── SUMMARY SHEET ────────────────────────────────────────────────────────────

def build_summary_sheet(wb, data, sub_names):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_view.showGridLines = False

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="Azure Environment Assessment")
    t.fill = TITLE_FILL; t.font = TITLE_FONT; t.alignment = CENTER
    ws.row_dimensions[1].height = 36

    now_str = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    subs_str = ", ".join(sub_names) if sub_names else "N/A"
    ws.merge_cells("A2:H2")
    s = ws.cell(row=2, column=1, value=f"Subscriptions: {subs_str}   |   Generated: {now_str}")
    s.font = Font(italic=True, name="Calibri", size=9, color="555555")
    s.alignment = CENTER
    ws.row_dimensions[2].height = 16

    # ── KPI tiles (row 4-6, columns A-H) ─────────────────────────────────────
    vms         = data.get("vms", [])
    disks       = data.get("disks", [])
    snapshots   = data.get("snapshots", [])
    sql         = data.get("sql", [])
    sql_mi_db   = data.get("sql_mi_db", [])
    sql_pools   = data.get("sql_pools", [])
    sql_vms     = data.get("sql_vms", [])
    storage     = data.get("storage", [])
    aks         = data.get("aks", [])
    functions   = data.get("functions", [])
    vaults      = data.get("backup_vaults", [])
    backup_items= data.get("backup_items", [])
    backup_sql  = data.get("backup_sql_items", [])
    cloud_spend = data.get("cloud_spend", [])

    total_resources = (len(vms) + len(disks) + len(snapshots) + len(sql) +
                       len(storage) + len(data.get("netapp", [])) +
                       len(data.get("cosmosdb", [])) + len(data.get("synapse", [])) +
                       len(aks) + len(data.get("aci", [])) + len(functions) +
                       len(data.get("avd", [])) + len(data.get("redis", [])))

    disk_gib  = sum(r.get("Size (GiB)", 0) for r in disks if isinstance(r.get("Size (GiB)"), (int, float)))
    snap_gib  = sum(r.get("Size (GiB)", 0) for r in snapshots if isinstance(r.get("Size (GiB)"), (int, float)))
    vm_gib    = sum(r.get("Total Storage (GiB)", 0) for r in vms if isinstance(r.get("Total Storage (GiB)"), (int, float)))
    stor_gib  = sum(r.get("Total Size (GiB)", 0) for r in storage if isinstance(r.get("Total Size (GiB)"), (int, float)))
    total_tib = round((disk_gib + snap_gib + stor_gib) / 1024, 2)

    running  = sum(1 for v in vms if "running" in str(v.get("Power State","")).lower())
    stopped  = len(vms) - running
    unattach = sum(1 for d in disks if "unattached" in str(d.get("Disk State","")).lower())

    # VM backup coverage
    protected_vm_names = {str(i.get("Protected Item","")).lower() for i in backup_items}
    vms_protected = sum(1 for v in vms if str(v.get("Name","")).lower() in protected_vm_names
                        or str(v.get("Backup Protected","")).lower() == "yes")
    if len(vms) > 0:
        bk_pct = round(vms_protected / len(vms) * 100)
        bk_label = f"VM BACKUP\n{bk_pct}% COVERED"
        bk_value = f"{vms_protected}/{len(vms)}"
    else:
        bk_label = "VM BACKUP\nCOVERAGE"
        bk_value = "N/A"

    # Monthly spend — sum all current-month cost columns across all spend rows
    cur_month_spend = 0.0
    spend_currency  = "USD"
    if cloud_spend:
        cur_keys = [k for k in cloud_spend[0] if k.startswith("Cost (")]
        if cur_keys:
            cur_month_spend = round(sum(r.get(cur_keys[0], 0) for r in cloud_spend
                                        if isinstance(r.get(cur_keys[0]), (int, float))), 2)
            spend_currency  = cloud_spend[0].get("Currency", "USD")

    spend_label = f"MONTHLY SPEND\n({spend_currency})" if cur_month_spend else "MONTHLY SPEND\n(N/A)"
    spend_value = f"${cur_month_spend:,.0f}" if cur_month_spend else "N/A"

    kpis = [
        ("TOTAL\nRESOURCES",  total_resources),
        ("TOTAL STORAGE\n(TiB)", total_tib),
        ("VMs\nRunning",     running),
        ("VMs\nStopped",     stopped),
        ("SQL\nDatabases",   len(sql)),
        ("Storage\nAccounts",len(storage)),
        (bk_label,           bk_value),
        (spend_label,        spend_value),
    ]

    ws.row_dimensions[3].height = 8
    for col, (label, value) in enumerate(kpis, 1):
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col)
        vr = ws.cell(row=4, column=col, value=value)
        vr.fill = KPI_FILL; vr.font = KPI_FONT; vr.alignment = CENTER
        lr = ws.cell(row=5, column=col, value=label)
        lr.fill = KPI_FILL; lr.font = KPI_LBL; lr.alignment = CENTER
        ws.cell(row=6, column=col).fill = KPI_FILL
    ws.row_dimensions[4].height = 38
    ws.row_dimensions[5].height = 26
    ws.row_dimensions[6].height = 6

    ws.row_dimensions[7].height = 8  # spacer

    # ── Two-column layout: Workload Inventory (left) | Risk & Findings (right) ─
    LEFT_COLS  = (1, 4)   # A-D
    RIGHT_COLS = (5, 8)   # E-H

    def section_header(row, col_start, col_end, title, fill=HEADER_FILL):
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=title)
        c.fill = fill; c.font = HEADER_FONT; c.alignment = CENTER
        ws.row_dimensions[row].height = 18
        return row + 1

    def mini_header(row, labels, col_start):
        for i, lbl in enumerate(labels):
            c = ws.cell(row=row, column=col_start + i, value=lbl)
            c.fill = PatternFill("solid", fgColor="2E75B6")
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
            c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[row].height = 16
        return row + 1

    # Left column — Workload Inventory
    r = 8
    r = section_header(r, 1, 4, "Workload Inventory")
    r = mini_header(r, ["Workload", "Count", "Storage (GiB)", "Storage (TiB)"], 1)

    def stor_gib_for(rows_, key="Size (GiB)"):
        return round(sum(x.get(key,0) for x in rows_ if isinstance(x.get(key),(int,float))), 2)

    def mi_storage_gib(rows_):
        # Instance Storage is repeated for every DB row on the same MI — sum once per instance
        seen = {}
        for x in rows_:
            inst = x.get("Managed Instance","")
            val  = x.get("Instance Storage (GiB)", 0)
            if inst and isinstance(val, (int, float)) and inst not in seen:
                seen[inst] = val
        return round(sum(seen.values()), 2)

    def _tib(g): return round(g / 1024, 4)

    workloads = [
        ("Virtual Machines",        vms,                          vm_gib),
        ("Managed Disks",           disks,                        stor_gib_for(disks,"Size (GiB)")),
        ("Disk Snapshots",          snapshots,                    stor_gib_for(snapshots,"Size (GiB)")),
        ("Azure SQL",               sql,                          0),
        ("SQL Managed Instances",   sql_mi_db,                    mi_storage_gib(sql_mi_db)),
        ("SQL Elastic Pools",       sql_pools,                    0),
        ("SQL Server VMs",          sql_vms,                      0),
        ("Storage Accounts",        storage,                      stor_gib),
        ("Azure File Shares",       data.get("file_shares",[]),   stor_gib_for(data.get("file_shares",[]),"Quota (GiB)")),
        ("Azure NetApp Files",      data.get("netapp",[]),        stor_gib_for(data.get("netapp",[]),"Quota (GiB)")),
        ("Cosmos DB",               data.get("cosmosdb",[]),      0),
        ("Synapse Analytics",       data.get("synapse",[]),       0),
        ("AKS Clusters",            aks,                          0),
        ("Container Instances",     data.get("aci",[]),           0),
        ("Function Apps",           functions,                    0),
        ("Azure Virtual Desktop",   data.get("avd",[]),           0),
        ("Redis Cache",             data.get("redis",[]),         0),
        ("Backup Vaults",           vaults,                       0),
    ]

    for i, (name, rows_, sg) in enumerate(workloads):
        alt  = i % 2 == 0
        fill = ALT_FILL if alt else None
        cnt  = len(rows_)
        for col_off, val in enumerate([name, cnt, round(sg,2), _tib(sg)]):
            c = ws.cell(row=r, column=1+col_off, value=val)
            c.font = NORMAL; c.alignment = LEFT; c.border = BORDER
            if fill: c.fill = fill
        ws.row_dimensions[r].height = 15
        r += 1

    # Right column — Risk & Findings
    # Compute risks
    public_storage  = sum(1 for s in storage if "true" in str(s.get("Public Blob Access","")).lower())
    http_storage    = sum(1 for s in storage if not s.get("HTTPS Only", True))
    unattached_disks= unattach
    public_sql      = sum(1 for s in sql if "enabled" in str(s.get("Public Access","")).lower()
                          or "true" in str(s.get("Public Access","")).lower())
    non_ssl_redis   = sum(1 for r_ in data.get("redis",[]) if r_.get("Non-SSL Port"))
    vms_no_backup   = max(0, len(vms) - sum(v.get("Protected Items",0) for v in vaults))

    no_blob_soft_delete  = sum(1 for s in storage if s.get("Blob Soft Delete (days)") == ""
                               and s.get("Active Services","") not in ("","Unknown")
                               and "Files" not in s.get("Service Type",""))
    no_file_soft_delete  = sum(1 for s in storage if s.get("File Soft Delete (days)") == ""
                               and ("Files" in s.get("Active Services","") or
                                    "Files" in s.get("Service Type","")))
    mssql_no_backup = sum(1 for v in vms
                          if v.get("MSSQL-INSTALLED") == "Yes"
                          and str(v.get("Backup Protected","")).lower() != "yes")

    findings = [
        ("CRITICAL", "Public Blob Access Enabled",              public_storage),
        ("CRITICAL", "SQL with Public Network Access",          public_sql),
        ("HIGH",     "Storage Without HTTPS-Only",              http_storage),
        ("HIGH",     "Unattached Managed Disks",                unattached_disks),
        ("HIGH",     "Redis with Non-SSL Port Enabled",         non_ssl_redis),
        ("HIGH",     "Blob Storage: No Soft Delete",            no_blob_soft_delete),
        ("HIGH",     "Azure Files: No Soft Delete",             no_file_soft_delete),
        ("HIGH",     "SQL Server VMs Without Backup Coverage",  mssql_no_backup),
        ("MEDIUM",   "VMs Without Backup Coverage",             vms_no_backup),
    ]

    r2 = 8
    r2 = section_header(r2, 5, 8, "Risk & Findings")
    r2 = mini_header(r2, ["Severity", "Finding", "", "Count"], 5)

    sev_fills = {"CRITICAL": CRIT_FILL, "HIGH": HIGH_FILL, "MEDIUM": MED_FILL}
    for severity, finding, count in findings:
        c_sev = ws.cell(row=r2, column=5, value=severity)
        c_sev.fill = sev_fills.get(severity, PatternFill())
        c_sev.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
        c_sev.alignment = CENTER; c_sev.border = BORDER
        c_find = ws.cell(row=r2, column=6, value=finding)
        c_find.font = NORMAL; c_find.alignment = LEFT; c_find.border = BORDER
        ws.merge_cells(start_row=r2, start_column=6, end_row=r2, end_column=7)
        c_cnt = ws.cell(row=r2, column=8, value=count)
        c_cnt.fill = (RED_FILL if count > 0 else GREEN_FILL)
        c_cnt.font = BOLD; c_cnt.alignment = CENTER; c_cnt.border = BORDER
        ws.row_dimensions[r2].height = 15
        r2 += 1

    # Azure Backup Infrastructure (right column, continued)
    r2 += 1
    r2 = section_header(r2, 5, 8, "Azure Backup Infrastructure")
    total_protected = sum(v.get("Protected Items",0) for v in vaults)
    backup_stats = [
        ("Recovery Services Vaults",  len(vaults)),
        ("VM Protected Items",        total_protected),
        ("SQL Protected Items",       len(backup_sql)),
    ]
    for k, v in backup_stats:
        ws.cell(row=r2, column=5, value=k).font = NORMAL
        ws.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=7)
        ws.cell(row=r2, column=5).border = BORDER
        ws.cell(row=r2, column=8, value=v).font = BOLD
        ws.cell(row=r2, column=8).alignment = CENTER
        ws.cell(row=r2, column=8).border = BORDER
        ws.row_dimensions[r2].height = 15
        r2 += 1

    # Region distribution (right column)
    r2 += 1
    r2 = section_header(r2, 5, 8, "Region Distribution")
    region_counter = defaultdict(int)
    for svc in data.values():
        if not isinstance(svc, list):
            continue
        for item in svc:
            loc = item.get("Location","")
            if loc:
                region_counter[loc] += 1
    for loc, cnt in sorted(region_counter.items(), key=lambda x: -x[1])[:10]:
        bar = "█" * min(cnt, 20)
        ws.cell(row=r2, column=5, value=loc).font = NORMAL
        ws.cell(row=r2, column=5).border = BORDER
        ws.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=7)
        ws.cell(row=r2, column=8, value=cnt).font = BOLD
        ws.cell(row=r2, column=8).alignment = CENTER
        ws.cell(row=r2, column=8).border = BORDER
        ws.row_dimensions[r2].height = 15
        r2 += 1

    # Storage by service
    r2 += 1
    r2 = section_header(r2, 5, 8, "Storage by Service (GiB)")
    storage_by_svc = [
        ("Storage Accounts",  stor_gib),
        ("Managed Disks",     stor_gib_for(disks, "Size (GiB)")),
        ("VMs (boot disks)",  vm_gib),
        ("Disk Snapshots",    stor_gib_for(snapshots, "Size (GiB)")),
        ("Azure File Shares", stor_gib_for(data.get("file_shares",[]), "Quota (GiB)")),
        ("Azure NetApp Files",stor_gib_for(data.get("netapp",[]), "Quota (GiB)")),
    ]
    storage_by_svc.sort(key=lambda x: x[1], reverse=True)
    for svc_name, sg in storage_by_svc:
        ws.cell(row=r2, column=5, value=svc_name).font = NORMAL
        ws.cell(row=r2, column=5).border = BORDER
        ws.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=7)
        ws.cell(row=r2, column=8, value=round(sg, 2)).font = BOLD
        ws.cell(row=r2, column=8).alignment = CENTER
        ws.cell(row=r2, column=8).border = BORDER
        ws.row_dimensions[r2].height = 15
        r2 += 1

    # Cloud Spend by Service (right column)
    if cloud_spend:
        r2 += 1
        cur_keys = [k for k in cloud_spend[0] if k.startswith("Cost (")]
        cur_key  = cur_keys[0] if cur_keys else None
        if cur_key:
            r2 = section_header(r2, 5, 8, f"Monthly Spend by Service  ({cur_key.replace('Cost (','').rstrip(')')})")
            top_spend = sorted(cloud_spend, key=lambda x: x.get(cur_key, 0), reverse=True)[:10]
            for item in top_spend:
                svc  = item.get("Service", "")
                cost = item.get(cur_key, 0)
                ws.cell(row=r2, column=5, value=svc).font = NORMAL
                ws.cell(row=r2, column=5).border = BORDER
                ws.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=7)
                c = ws.cell(row=r2, column=8, value=f"${cost:,.2f}")
                c.font = BOLD; c.alignment = CENTER; c.border = BORDER
                ws.row_dimensions[r2].height = 15
                r2 += 1
            # Total row
            total_cur = sum(x.get(cur_key, 0) for x in cloud_spend
                            if isinstance(x.get(cur_key), (int, float)))
            ws.cell(row=r2, column=5, value="TOTAL").font = Font(bold=True, name="Calibri", size=9)
            ws.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=7)
            for col in range(5, 9):
                ws.cell(row=r2, column=col).fill = HEADER_FILL
                ws.cell(row=r2, column=col).border = BORDER
            c = ws.cell(row=r2, column=8, value=f"${total_cur:,.2f}")
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
            c.alignment = CENTER
            ws.row_dimensions[r2].height = 16
            r2 += 1

    # ── Backup Sizing Summary (left column, below workload inventory) ──────────
    r += 1
    r = section_header(r, 1, 4, "Backup Sizing Summary", fill=PatternFill("solid", fgColor="375623"))
    r = mini_header(r, ["Service", "Total (GiB)", "Total (TiB)", "Suggested Method"], 1)

    blob_gib_total = stor_gib_for(storage, "Blob Size (GiB)")
    file_gib_total = stor_gib_for(storage, "File Size (GiB)")
    disk_gib_total = stor_gib_for(disks, "Size (GiB)")
    anf_gib_total  = stor_gib_for(data.get("netapp",[]), "Quota (GiB)")
    fs_gib_total   = stor_gib_for(data.get("file_shares",[]), "Quota (GiB)")
    mi_gib_total   = mi_storage_gib(sql_mi_db)

    sizing = [
        ("Managed Disks",          disk_gib_total,              "Azure Backup for Disks / Snapshots"),
        ("Blob Storage",           blob_gib_total,              "Azure Backup for Blobs / Versioning"),
        ("Azure Files",            file_gib_total + fs_gib_total, "Azure Backup for Files / File Sync"),
        ("Azure NetApp Files",     anf_gib_total,               "ANF Snapshots / CRR"),
        ("Azure SQL",              0,                           "Azure Backup for SQL / Auto-backup"),
        ("SQL Managed Instances",  mi_gib_total,                "Azure Backup for SQL MI / Auto-backup"),
    ]
    grand_gib = sum(g for _, g, _ in sizing)

    for i, (svc, sg, method) in enumerate(sizing):
        alt  = i % 2 == 0
        fill = ALT_FILL if alt else None
        cells = [svc, round(sg,2), _tib(sg), method]
        for col_off, val in enumerate(cells):
            c = ws.cell(row=r, column=1+col_off, value=val)
            c.font = NORMAL; c.alignment = LEFT; c.border = BORDER
            if fill: c.fill = fill
        ws.row_dimensions[r].height = 15
        r += 1

    # Grand total row
    ws.cell(row=r, column=1, value="TOTAL PROTECTABLE").font = BOLD
    ws.cell(row=r, column=2, value=round(grand_gib,2)).font = BOLD
    ws.cell(row=r, column=3, value=_tib(grand_gib)).font = BOLD
    ws.cell(row=r, column=4, value="").font = BOLD
    for col in range(1,5):
        ws.cell(row=r, column=col).fill = HEADER_FILL
        ws.cell(row=r, column=col).font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        ws.cell(row=r, column=col).border = BORDER
        ws.cell(row=r, column=col).alignment = LEFT
    ws.row_dimensions[r].height = 18
    r += 1

    # ── Snapshot Coverage Summary (left column) ───────────────────────────────
    r += 1
    r = section_header(r, 1, 4, "Disk Snapshot Coverage")
    snap_no      = sum(1 for d in disks if "No Snapshot" in str(d.get("Snapshot Coverage","")))
    snap_stale   = sum(1 for d in disks if "Stale"       in str(d.get("Snapshot Coverage","")))
    snap_aging   = sum(1 for d in disks if "Aging"       in str(d.get("Snapshot Coverage","")))
    snap_current = sum(1 for d in disks if "Current" in str(d.get("Snapshot Coverage","")) or
                                           "Recent"  in str(d.get("Snapshot Coverage","")))
    snap_rows = [
        ("No Snapshot",       snap_no,      RED_FILL),
        ("Stale (>30 days)",  snap_stale,   RED_FILL),
        ("Aging (8–30 days)", snap_aging,   YELLOW_FILL),
        ("Current (≤7 days)", snap_current, GREEN_FILL),
    ]
    for label, cnt, cfill in snap_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=1, value=label).font = NORMAL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = LEFT
        c = ws.cell(row=r, column=4, value=cnt)
        c.font = BOLD; c.fill = cfill; c.border = BORDER; c.alignment = CENTER
        ws.row_dimensions[r].height = 15
        r += 1

    # Add snapshot warning to findings if significant
    snap_unprotected_gib = sum(
        d.get("Size (GiB)",0) for d in disks
        if isinstance(d.get("Size (GiB)"),(int,float)) and
        ("No Snapshot" in str(d.get("Snapshot Coverage","")) or
         "Stale"       in str(d.get("Snapshot Coverage","")))
    )

    # Column widths — balanced for KPI tiles (one per col) and body sections
    # Left body: A=workload name, B=count, C=GiB, D=TiB
    # Right body: E=severity, F+G=finding text (merged), H=count
    col_w = [24, 11, 14, 12, 13, 24, 20, 14]
    for col, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return ws


# ─── WORKBOOK ─────────────────────────────────────────────────────────────────

def _add_snapshot_coverage(data):
    """Cross-reference disks with snapshots; stamp each disk with coverage status."""
    snap_ages = defaultdict(list)
    for snap in data.get("snapshots", []):
        src  = snap.get("Source Disk", "")
        age  = snap.get("Age (days)", "")
        if src and isinstance(age, int):
            snap_ages[src].append(age)

    for disk in data.get("disks", []):
        name = disk.get("Name", "")
        ages = snap_ages.get(name, [])
        if not ages:
            disk["Snapshot Coverage"] = "No Snapshot"
        else:
            newest = min(ages)
            if newest <= 1:
                disk["Snapshot Coverage"] = f"Current ({newest}d)"
            elif newest <= 7:
                disk["Snapshot Coverage"] = f"Recent ({newest}d)"
            elif newest <= 30:
                disk["Snapshot Coverage"] = f"Aging ({newest}d)"
            else:
                disk["Snapshot Coverage"] = f"Stale ({newest}d)"


def build_scenario_builder(data, output_path):
    """
    Write a Veeam Scenario Builder import file (CAzureWrapper table format).

    Loads the bundled template (veeam_scenario_builder_template.xlsx) to
    preserve the exact table structure, data validations, column widths, and
    table style that Scenario Builder expects, then fills in one row per
    workload type sized from the collected assessment data.

    Default change rates / compression are conservative starting points the
    user can tune inside the Scenario Builder UI.
    """
    import zipfile, io as _io, os as _os2, re as _re

    def _gib_sum(rows, key):
        return sum(r.get(key, 0) or 0 for r in rows if isinstance(r.get(key), (int, float)))

    def _mi_gib(rows):
        # Instance Storage is repeated per-DB row — sum once per managed instance
        seen = {}
        for r in rows:
            inst = r.get("Managed Instance", "")
            val  = r.get("Instance Storage (GiB)", 0)
            if inst and isinstance(val, (int, float)) and inst not in seen:
                seen[inst] = val
        return sum(seen.values())

    def _tib(gib):
        return round(gib / 1024, 4) if gib else 0

    # ── Compute per-workload totals ───────────────────────────────────────────
    vms         = data.get("vms", [])
    disks       = data.get("disks", [])
    sql         = data.get("sql", [])
    sql_mi      = data.get("sql_mi_db", [])
    sql_pools   = data.get("sql_pools", [])
    storage     = data.get("storage", [])
    file_shares = data.get("file_shares", [])
    netapp      = data.get("netapp", [])
    cosmosdb    = data.get("cosmosdb", [])
    synapse     = data.get("synapse", [])

    # VM storage: prefer disk inventory (more accurate), fall back to VM rows
    vm_gib = _gib_sum(disks, "Size (GiB)")
    if not vm_gib:
        vm_gib = _gib_sum(vms, "Total Storage (GiB)")

    sql_gib      = _gib_sum(sql,       "Allocated (GiB)") or _gib_sum(sql,      "Used (GiB)")
    sql_mi_gib   = _mi_gib(sql_mi)
    sql_pool_gib = _gib_sum(sql_pools, "Max Storage (GiB)") or _gib_sum(sql_pools, "Allocated (GiB)")

    blob_gib    = _gib_sum(storage,     "Blob Size (GiB)") or _gib_sum(storage, "Total Size (GiB)")
    file_sa_gib = _gib_sum(storage,     "File Size (GiB)")
    file_sh_gib = _gib_sum(file_shares, "Used Size (GiB)")
    anf_gib     = _gib_sum(netapp,      "Used Size (GiB)") or _gib_sum(netapp, "Quota (GiB)")
    cosmos_gib  = _gib_sum(cosmosdb,    "Data (GiB)") + _gib_sum(cosmosdb, "Index (GiB)")
    synapse_gib = _gib_sum(synapse,     "Used (GiB)")

    # ── Build workload rows ───────────────────────────────────────────────────
    # Columns: Workload name | Instance count | VM data (TB) | VM chg% |
    #          DB data (TB)  | DB chg%        | File data (TB)| File chg% |
    #          Compress%     | Growth%
    #
    # Change rate defaults:  VMs 5% | Databases 10% | Files/blob 3%
    # Compression: 30%  |  Growth: 10% annual

    rows = []
    if vms or disks:
        rows.append(("Virtual Machines",        len(vms),         _tib(vm_gib),       5,  0, 0,               0, 0, 30, 10))
    if sql:
        rows.append(("Azure SQL Databases",     len(sql),         0, 0, _tib(sql_gib),     10,               0, 0, 30, 10))
    if sql_mi:
        rows.append(("SQL Managed Instances",   len(sql_mi),      0, 0, _tib(sql_mi_gib),  10,               0, 0, 30, 10))
    if sql_pools:
        rows.append(("SQL Elastic Pools",       len(sql_pools),   0, 0, _tib(sql_pool_gib),10,               0, 0, 30, 10))
    if blob_gib:
        rows.append(("Azure Blob Storage",      len(storage),     0, 0,               0, 0, _tib(blob_gib),   3, 30, 10))
    if file_sa_gib:
        rows.append(("Azure Files (Storage)",   len(storage),     0, 0,               0, 0, _tib(file_sa_gib), 3, 30, 10))
    if file_shares:
        rows.append(("Azure File Shares",       len(file_shares), 0, 0,               0, 0, _tib(file_sh_gib), 3, 30, 10))
    if netapp:
        rows.append(("Azure NetApp Files",      len(netapp),      0, 0,               0, 0, _tib(anf_gib),    3, 30, 10))
    if cosmosdb:
        rows.append(("Azure Cosmos DB",         len(cosmosdb),    0, 0, _tib(cosmos_gib),  5,               0, 0, 30, 10))
    if synapse:
        rows.append(("Azure Synapse Analytics", len(synapse),     0, 0, _tib(synapse_gib), 10,               0, 0, 30, 10))

    # ── Load template (preserves table structure, validations, style) ─────────
    template_path = _os2.path.join(_os2.path.dirname(_os2.path.abspath(__file__)),
                                   "veeam_scenario_builder_template.xlsx")
    if not _os2.path.exists(template_path):
        raise FileNotFoundError(
            f"Veeam Scenario Builder template not found: {template_path}\n"
            "Ensure veeam_scenario_builder_template.xlsx is in the same folder as azure_assessment.py"
        )

    with open(template_path, "rb") as f:
        tmpl_bytes = f.read()

    wb = openpyxl.load_workbook(_io.BytesIO(tmpl_bytes))
    ws = wb["Azure"]

    # Clear existing data rows (rows 2 → 51; row 52 is the totals row)
    for row_idx in range(2, 52):
        for col_idx in range(1, 11):
            ws.cell(row=row_idx, column=col_idx).value = None

    # Write our workload rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = val

    # Save to a buffer first, then patch app.xml back to the template's version.
    # openpyxl stamps its own Application string; the Veeam Scenario Builder
    # web validator rejects anything that isn't "Microsoft Excel".
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Extract the original app.xml from the template
    with zipfile.ZipFile(_io.BytesIO(tmpl_bytes)) as zt:
        orig_app_xml = zt.read("docProps/app.xml")

    # Rewrite the zip, swapping in the original app.xml
    buf.seek(0)
    out_buf = _io.BytesIO()
    with zipfile.ZipFile(buf, "r") as zin, zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == "docProps/app.xml":
                zout.writestr(item, orig_app_xml)
            else:
                zout.writestr(item, zin.read(item.filename))

    with open(output_path, "wb") as f:
        f.write(out_buf.getvalue())


def build_workbook(data, output_path, sub_names, anonymizer=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Post-processing: stamp snapshot coverage onto each disk
    _add_snapshot_coverage(data)

    # Apply anonymization to all collected rows before writing
    if anonymizer:
        data = {k: anonymizer.apply(v) for k, v in data.items()}

    build_summary_sheet(wb, data, sub_names)
    build_sheet_vms(wb,            data.get("vms", []))
    build_sheet_disks(wb,          data.get("disks", []))
    build_sheet_snapshots(wb,      data.get("snapshots", []))
    build_sheet_sql(wb,            data.get("sql", []))
    build_sheet_sql_mi_databases(wb, data.get("sql_mi_db", []))
    build_sheet_elastic_pools(wb,  data.get("sql_pools", []))
    build_sheet_sql_vms(wb,        data.get("sql_vms", []))
    build_sheet_storage(wb,        data.get("storage", []))
    build_sheet_file_shares(wb,    data.get("file_shares", []))
    build_sheet_netapp(wb,         data.get("netapp", []))
    build_sheet_cosmosdb(wb,       data.get("cosmosdb", []))
    build_sheet_synapse(wb,        data.get("synapse", []))
    build_sheet_aks(wb,            data.get("aks", []))
    build_sheet_aci(wb,            data.get("aci", []))
    build_sheet_functions(wb,      data.get("functions", []))
    build_sheet_avd(wb,            data.get("avd", []))
    build_sheet_redis(wb,          data.get("redis", []))
    build_sheet_backup_vaults(wb,       data.get("backup_vaults", []))
    build_sheet_backup_items(wb,        data.get("backup_items", []))
    build_sheet_backup_sql_items(wb,    data.get("backup_sql_items", []))
    build_sheet_backup_policies(wb,     data.get("backup_policies", []))
    build_sheet_backup_costs(wb,        data.get("backup_costs", []))
    build_sheet_cloud_spend(wb,    data.get("cloud_spend", []))

    wb.save(output_path)
    log.info("Saved: %s", output_path)


# ─── SUBSCRIPTION COLLECTOR ───────────────────────────────────────────────────

def collect_subscription(credential, sub_id, sub_name, args):
    """Run all collectors for a single subscription."""
    verbose = args.verbose
    log.info("Scanning subscription: %s (%s)", sub_name, sub_id)

    collectors = [
        ("vms",            lambda: collect_vms(credential, sub_id, sub_name, verbose)),
        ("disks",          lambda: collect_disks(credential, sub_id, sub_name, verbose)),
        ("sql",            lambda: collect_sql(credential, sub_id, sub_name, verbose)),
        ("sql_mi_db",      lambda: collect_sql_mi_databases(credential, sub_id, sub_name, verbose)),
        ("sql_pools",      lambda: collect_sql_elastic_pools(credential, sub_id, sub_name, verbose)),
        ("storage",        lambda: collect_storage(credential, sub_id, sub_name, verbose)),
        ("file_shares",    lambda: collect_file_shares(credential, sub_id, sub_name, verbose)),
        ("cosmosdb",       lambda: collect_cosmosdb(credential, sub_id, sub_name, verbose)),
        ("aks",            lambda: collect_aks(credential, sub_id, sub_name, verbose)),
        ("aci",            lambda: collect_container_instances(credential, sub_id, sub_name, verbose)),
        ("functions",      lambda: collect_functions(credential, sub_id, sub_name, verbose)),
        ("redis",          lambda: collect_redis(credential, sub_id, sub_name, verbose)),
        ("netapp",         lambda: collect_netapp(credential, sub_id, sub_name, verbose)),
        ("avd",            lambda: collect_avd(credential, sub_id, sub_name, verbose)),
        ("synapse",        lambda: collect_synapse(credential, sub_id, sub_name, verbose)),
    ]

    if not args.skip_snapshots:
        collectors.insert(2, ("snapshots", lambda: collect_snapshots(credential, sub_id, sub_name, verbose)))

    results = {}
    for name, fn in tqdm(collectors, desc=f"  {sub_name[:30]}", leave=False):
        try:
            results[name] = fn()
        except Exception as exc:
            log.warning("Collector '%s' failed for %s: %s", name, sub_name, exc)
            results[name] = []

    # SQL VMs (returns tuple: rows + name set)
    sqlvm_rows, sql_vm_names = collect_sql_vms(credential, sub_id, sub_name, verbose)
    results["sql_vms"] = sqlvm_rows

    # Backup (returns tuple: vault rows + item rows)
    vault_rows, item_rows = collect_backup(credential, sub_id, sub_name, verbose)
    results["backup_vaults"] = vault_rows
    results["backup_items"]  = item_rows

    # SQL workload backup items and backup policies
    results["backup_sql_items"]  = collect_backup_sql_items(credential, sub_id, sub_name, verbose)
    results["backup_policies"]   = collect_backup_policies(credential, sub_id, sub_name, verbose)

    # Cost Management data
    results["backup_costs"]  = collect_backup_costs(credential, sub_id, sub_name, verbose)
    results["cloud_spend"]   = collect_cloud_spend(credential, sub_id, sub_name, verbose)

    # Cross-reference: stamp SQL Server flag and backup policy onto each VM row
    if sql_vm_names or item_rows:
        # Build lookup: vm_name_lower -> policy_name (first match wins)
        backup_policy_map = {}
        for itm in item_rows:
            item_name = itm.get("Protected Item", "").lower()
            policy    = itm.get("Policy Name", "")
            if item_name and item_name not in backup_policy_map:
                backup_policy_map[item_name] = policy

        for vm_row in results.get("vms", []):
            vm_lower = vm_row["Name"].lower()
            if vm_lower in sql_vm_names:
                vm_row["MSSQL-INSTALLED"] = "Yes"
            policy = backup_policy_map.get(vm_lower, "")
            if policy:
                vm_row["Backup Policy"]    = policy
                vm_row["Backup Protected"] = "Yes"
            elif not vm_row["Backup Policy"]:
                vm_row["Backup Protected"] = "No"

    return results


# ─── CLI ──────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description=(
            "Azure Environment Assessment Tool — read-only inventory scanner that "
            "produces a multi-sheet Excel workbook covering every major Azure workload type."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument(
        "--subscription", nargs="+", metavar="SUB_ID",
        help="One or more Azure subscription IDs to scan (default: current subscription)",
    )
    p.add_argument(
        "--all-subscriptions", action="store_true",
        help="Scan all subscriptions accessible with the current credentials",
    )
    p.add_argument(
        "--tenant", metavar="TENANT_ID",
        help="Azure tenant ID (optional; useful for multi-tenant environments)",
    )
    p.add_argument(
        "--output", metavar="FILENAME",
        help="Output .xlsx filename (default: azure_assessment_<date>.xlsx)",
    )
    p.add_argument(
        "--workers", type=int, default=4, metavar="N",
        help="Number of subscriptions scanned in parallel (default: 4)",
    )
    p.add_argument(
        "--skip-snapshots", action="store_true",
        help="Skip disk snapshot enumeration (faster on accounts with many snapshots)",
    )
    p.add_argument(
        "--verbose", action="store_true",
        help="Enable detailed per-service logging",
    )
    p.add_argument(
        "--anonymize", action="store_true",
        help="Replace resource names with opaque codes; saves a reversible mapping CSV alongside the workbook",
    )
    p.add_argument(
        "--scenario-builder", action="store_true",
        help="Also write a Veeam Scenario Builder import file (CAzureWrapper format) alongside the main workbook",
    )
    return p.parse_args()


def get_subscriptions(credential, args):
    """Return list of (sub_id, sub_name) tuples to scan."""
    if _SubscriptionClient is None:
        print("ERROR: azure-mgmt-resource not installed.")
        sys.exit(1)

    sub_client = _SubscriptionClient(credential)

    if args.subscription:
        results = []
        for sid in args.subscription:
            try:
                sub = sub_client.subscriptions.get(sid)
                results.append((sub.subscription_id, sub.display_name))
            except Exception:
                results.append((sid, sid))
        return results

    if args.all_subscriptions:
        return [(s.subscription_id, s.display_name)
                for s in safe_list(sub_client.subscriptions.list())
                if s.state and s.state.lower() == "enabled"]

    # Default: current subscription from env var or first accessible
    env_sub = _os.environ.get("AZURE_SUBSCRIPTION_ID", "")
    if env_sub:
        try:
            sub = sub_client.subscriptions.get(env_sub)
            return [(sub.subscription_id, sub.display_name)]
        except Exception:
            return [(env_sub, env_sub)]

    # Fall back to first accessible subscription
    all_subs = [(s.subscription_id, s.display_name)
                for s in safe_list(sub_client.subscriptions.list())
                if s.state and s.state.lower() == "enabled"]
    if not all_subs:
        print("No accessible subscriptions found. Run: az login")
        sys.exit(1)
    return [all_subs[0]]


def main():
    args = parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.WARNING,
        format="%(levelname)s  %(message)s",
    )

    # Ensure az CLI is findable regardless of which Python interpreter launched us.
    # Homebrew and some conda setups don't add /opt/homebrew/bin to the subprocess PATH.
    _extra_paths = ["/opt/homebrew/bin", "/usr/local/bin", "/home/linuxbrew/.linuxbrew/bin"]
    _env_path = _os.environ.get("PATH", "")
    for _p in _extra_paths:
        if _p not in _env_path:
            _os.environ["PATH"] = _p + _os.pathsep + _env_path
            _env_path = _os.environ["PATH"]

    # Output filename
    date_str   = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%d")
    raw_output = args.output or f"azure_assessment_{date_str}.xlsx"
    output     = _os.path.basename(raw_output) if args.output else raw_output
    if not output.endswith(".xlsx"):
        output += ".xlsx"

    print("Azure Environment Assessment Tool")
    print("=" * 50)

    # Authenticate
    kwargs = {}
    if args.tenant:
        kwargs["tenant_id"] = args.tenant
    credential = DefaultAzureCredential(**kwargs)

    # Resolve subscriptions
    subscriptions = get_subscriptions(credential, args)
    print(f"Subscriptions to scan: {len(subscriptions)}")
    for sid, name in subscriptions:
        print(f"  {name} ({sid})")
    print()

    # Collect — parallel across subscriptions
    all_data    = defaultdict(list)
    sub_names   = []

    def _scan(sub_tuple):
        sid, sname = sub_tuple
        return sname, collect_subscription(credential, sid, sname, args)

    with ThreadPoolExecutor(max_workers=min(args.workers, len(subscriptions))) as exe:
        futures = {exe.submit(_scan, s): s for s in subscriptions}
        for fut in tqdm(as_completed(futures), total=len(futures), desc="Subscriptions"):
            try:
                sname, results = fut.result()
                sub_names.append(sname)
                for key, rows in results.items():
                    all_data[key].extend(rows)
            except Exception as exc:
                sub_id = futures[fut][0]
                log.warning("Subscription %s failed: %s", sub_id, exc)

    # Build workbook
    print(f"\nBuilding workbook: {output}")
    anonymizer = Anonymizer() if args.anonymize else None
    build_workbook(dict(all_data), output, sub_names, anonymizer=anonymizer)

    if anonymizer:
        mapping_path = output.replace(".xlsx", "_mapping.csv")
        anonymizer.save(mapping_path)
        print(f"Anonymization mapping : {mapping_path}  (keep this file private)")

    if getattr(args, "scenario_builder", False):
        sb_path = output.replace(".xlsx", "_scenario_builder.xlsx")
        print(f"Building Scenario Builder file: {sb_path}")
        build_scenario_builder(dict(all_data), sb_path)
        print(f"Scenario Builder file : {sb_path}")

    # Print summary
    print("\n── Results ───────────────────────────────────────")
    service_map = [
        ("Virtual Machines",      "vms"),
        ("Managed Disks",         "disks"),
        ("Disk Snapshots",        "snapshots"),
        ("Azure SQL",             "sql"),
        ("SQL MI Databases",      "sql_mi_db"),
        ("Storage Accounts",      "storage"),
        ("Azure File Shares",     "file_shares"),
        ("Azure NetApp Files",    "netapp"),
        ("Cosmos DB",             "cosmosdb"),
        ("Synapse Analytics",     "synapse"),
        ("AKS Clusters",          "aks"),
        ("Container Instances",   "aci"),
        ("Function Apps",         "functions"),
        ("Azure Virtual Desktop", "avd"),
        ("Redis Cache",           "redis"),
        ("Backup Vaults",         "backup_vaults"),
        ("Backup Protected Items","backup_items"),
    ]
    for label, key in service_map:
        cnt = len(all_data.get(key, []))
        if cnt:
            print(f"  {label:<28} {cnt:>6}")

    total_disk_gib = sum(
        r.get("Size (GiB)", 0) for r in all_data.get("disks", [])
        if isinstance(r.get("Size (GiB)"), (int, float))
    )
    total_stor_gib = sum(
        r.get("Total Size (GiB)", 0) for r in all_data.get("storage", [])
        if isinstance(r.get("Total Size (GiB)"), (int, float))
    )
    print(f"\n  Managed Disk Storage:  {round(total_disk_gib, 2):>8} GiB  "
          f"({round(total_disk_gib/1024, 3)} TiB)")
    print(f"  Blob/File Storage:     {round(total_stor_gib, 2):>8} GiB  "
          f"({round(total_stor_gib/1024, 3)} TiB)")
    print(f"\nOutput: {output}")


if __name__ == "__main__":
    main()
